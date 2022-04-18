"""

AUTOMATION CONTROL CODE - FULL CONTROL - V5 - 2021

Atlantic Cancer Research Institute - ACRI

Code originally written by Andre LeBlanc, electrical engineering student at l'Université de Moncton
August 2021

NOTE 1: ALL USB cables need to be connected to the host ENA prior to running the code. Otherwise code will show an error.
NOTE 2: Prior to running the code, assure that the switchs' USB driver is in its correct format (lib32). The driver can be configured with Zadig.exe
"""

import pyvisa
# from Phidget22.Phidget import *
from Phidget22.Devices.VoltageOutput import *
import PySimpleGUI as sg
from pymodbus.client.sync import ModbusSerialClient as ModbusClient
# import relay_ft245r
import sys
import os
import time
import datetime
import pytz
import statistics
import openpyxl

"""
************************************************
**************** GLOBAL OBJECTS ****************
************************************************
"""

"""
SWITCH

# rb = relay_ft245r.FT245R()
# dev_list = rb.list_dev()
"""

"""
NETWORK ANALYZER
"""
# Create a ressource manager
rm = pyvisa.ResourceManager()
rm.list_resources()

# Open the Network analyzer by name
analyzer = rm.open_resource('USB0::0x2A8D::0x0001::MY55201231::0::INSTR')

"""
GENERATOR
"""
UNIT = 0x01  # Set the generator slave address

# COM varies from generator to generator
client = ModbusClient(method='rtu', port='COM7', timeout=4, baudrate=115200, strict=False)


"""
*******************************************
**************** FUNCTIONS ****************
*******************************************
"""

"""
INITS
"""
def init():
    print("\n------------------CODE INITIALISATIONS------------------\n\n")
    """
    1. GENERATOR INIT
    """
    client.set_parity = 0
    client.set_bytesize = 8
    client.set_stopbits = 1
    client.connect()
    print("Generator connected \n")

    # TIMEOUT ( Must be greater than Power_ON_time !!!)
    client.write_register(98, 3000, unit=UNIT)

    # RESET FAULT : set bit 7 of register 2 to 1 and then to 0
    # (see page 19 of Microwave generator documentation --> Cabinet 18, room 455, IARC)
    client.write_register(2, 0x80, unit=UNIT)

    # RESET FAULT: set bit 7 of register 2 back to 0
    # TURN GENERATOR OFF (default state) : set register 2, bit 6 to
    client.write_register(2, 0x00, unit=UNIT)
    print("Faults reseted\n")
    print("Microwave is turned off for its initial state \n")

    # GENERATOR STARTING MODE TO NORMAL : set registor 3 bit 0 to 0
    client.write_register(3, 0x00, unit=UNIT)

    """
    2. SWITCH INIT
    
    switch1 = 2
    switch2 = 3

    # list of FT245R devices are returned
    if len(dev_list) == 0:
        print('No FT245R devices found')
        sys.exit()

    # Show their serial numbers
    for dev in dev_list:
        print(dev.serial_number)

    # Pick the first one for simplicity
    dev = dev_list[0]
    # print('Using device with serial number ' + str(dev.serial_number))
    rb.connect(dev)

    # Place switch in state I for its initial state
    rb.switchon(switch1)
    rb.switchoff(switch2)

    """

    """
    3. NETWORK ANALYZER INIT
    """
    # Set the time for 10s
    analyzer.timeout = 15000  # essayer des delay entre les commande

    # Return the Analyzer'ID string to tell us it's connected
    print(analyzer.query('*IDN?'))

    # CHANNEL 1 INITIALISATION
    analyzer.write(":CALCulate1:PARameter:COUNt 1")
    analyzer.write("calculate1:measure1:format smith")
    analyzer.write("CALCulate1:MEASure1:PARameter 'S11'")  # Define the parameter for each trace = s11 pour le channel 1
    # analyzer.write("CALC:PAR:SEL 'CH1_S11_1,S11'")


"""
NORMAL TEST FUNCTION
"""
def Normal_Test(startFreq, stopFreq, datapoints, BW, num_its, ONdelay, OFFdelay, directory, freq,
                power, rpower, Peris_ON, RPM, Iso_ON, IsLog):
    print("------------------NORMAL TEST------------------\n")
    time.sleep(2)
    print("The test will begin in 3 seconds\n")
    time.sleep(3)

    # Update next step text color on GUI window
    window.find_element('_NORMAL2_').update(text_color='red')
    sg.OneLineProgressMeter('Test progress...', 1, 2 * num_its + 4, key='METER1', grab_anywhere=True)

    """
    1. NETWORK ANALYZER SET
    """
    # CONVERSION DE FRÉQUENCE DE MHZ À HZ
    startFreqHz = startFreq * 1000000
    stopFreqHz = stopFreq * 1000000
    # Set channel 1 freq and numpoint
    analyzer.write("SENS1:SWE:POIN " + str(datapoints))
    # analyzer.write("SENS1:SWE:POW " + str(InputPower))
    analyzer.write("SENS1:FREQ:START " + str(startFreqHz))
    analyzer.write("SENS1:FREQ:STOP " + str(stopFreqHz))
    analyzer.write("SENS1:BAND " + str(BW))

    # Determine, i.e. query, number of points in trace for ASCII transfer - query
    # Verification of data transfer between host PC and vector analyzer
    analyzer.write("SENS1:SWE:POIN?")
    numPoints = analyzer.read()
    print("Number of trace points (Channel 1)\n" + numPoints)

    # Determine, i.e. query, start and stop frequencies, i.e. stimulus begin and end points
    # Verification of data transfer between host PC and vector analyzer
    # analyzer.write("SENS1:FOM:RANG:FREQ:STAR?")
    analyzer.write("SENS1:FREQ:START?")
    startFreq = analyzer.read()
    analyzer.write("SENS1:FREQ:STOP?")
    stopFreq = analyzer.read()
    print("Analyzer start frequency (Channel 1) = \n" + startFreq + "\n" + "Stop frequency (Channel 1) = \n" + stopFreq)

    # SET FREQUENCY SWEEP TIME (MINIMUM DELAY FOR REAL TIME APPLICATION)
    analyzer.write("SENS1:SWE:TIME " + str(OFFdelay))

    # CHANNEL 1 PUT ON HOLD MODE FOR ITS DEFAULT TRIGGER STATE
    analyzer.write("SENS1:SWE:MODE HOLD")

    # Saved format specification (RI is selected)
    analyzer.write("MMEMory:STOR:TRAC:FORM:SNP RI")

    # Selection of sweep type (linear or logarithmic)
    if IsLog == 0:
        analyzer.write("SENS1:SWEep:TYPE LIN")
        print("Linear frequency sweep selected\n")

    if IsLog == 1:
        analyzer.write("SENS1:SWEep:TYPE LOG")
        print("Logarithmic frequency sweep selected\n")

    filename = directory

    try:
        analyzer.write("mmemory:mdirectory 'D:/" + filename + "'")
        print("D: drive folder created \n")

    except:
        print("D: folders couldn't be created.\n")

    """
    2. GENERATOR SET
    """
    # RESET COMMUNICATION TIMEOUT ( Must be greater than Power_ON_time !!!)
    # Function : register 98 set to X time before generator is faulted (red light on generator)
    client.write_register(98, 3000, unit=UNIT)

    # SET FREQUENCY
    # See page 31 of microwave generator user manual for more information (Cabinet 18, room 455, IARC)
    freqKHz = freq * 10
    client.write_register(9, freqKHz, unit=UNIT)
    rr = client.read_holding_registers(112, 1, unit=UNIT)
    rr1 = rr.registers
    print("Generator frequency is set to :" + str(rr1) + " Hz \n")

    # REFLECTED POWER LIMITATION MODE - ON
    client.write_register(2, 0x10, unit=UNIT)
    print("Reflected power limitation mode activated \n")

    # REFLECTED POWER SET
    # Value = 15% of transmitted power OR Manual selected value (See generator user manual for more details, cabinet 18, room 455, IARC)
    auto_rpower = 0.15 * power
    if int(rpower) != 0:
        client.write_register(1, int(rpower), unit=UNIT)
        rr = client.read_holding_registers(1, 1, unit=UNIT)
        rr1 = rr.registers
        print("Reflected power set to " + str(rr1) + " W \n")
    else:
        client.write_register(1, int(auto_rpower), unit=UNIT)
        rr = client.read_holding_registers(1, 1, unit=UNIT)
        rr1 = rr.registers
        print("Reflected power set to " + str(rr1) + " W \n")

    # TRANSMITTED POWER SET
    client.write_register(0, power, unit=UNIT)
    print("Transmitted power value set \n")

    # TURN GENERATOR OFF (default state)
    client.write_register(2, 0x00, unit=UNIT)
    print("Microwave is turned off for its initial state \n")

    """
    3. Peristaltic pump set
    """
    if Peris_ON == 1:
        ao4 = LucidControlAO4('COM6')

        # Open AO4 port
        if ao4.open() == False:
            print('Error connecting to port {0} '.format(
                ao4.portName))
            ao4.close()
            exit()

        # Create a tuple of 4 voltage objects
        values = (ValueVOS4(), ValueVOS4(), ValueVOS4(), ValueVOS4())

        # Initialize a boolean tuple for channels to change.
        channels = (True, True, True, True)

    """
    4. Isocratic pump set
    """
    if Iso_ON == 1:
        sg.popup(
            "Please wait 40 seconds for isocratic pump to initialise. Once done, enter desired flow speed in command prompt. Press Okay to continue.")
        os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY.exe")
        # window.findElement('_TEST_FRAME_').update(Visible=True)

    """
    TEST RUN
    """
    print("---------------------START TEST---------------------")

    for i in range(num_its + 1):
        if i == 0:

            # ------------- DUMP FIRST MEASUREMENT - SWITCH INIT -------------#
            # Update next step text color on GUI window
            window.find_element('_NORMAL2_').update(text_color='black')
            window.find_element('_NORMAL3_').update(text_color='red')
            sg.OneLineProgressMeter('Test progress...', 2, 2 * num_its + 4, key='METER1', grab_anywhere=True)

            # INITIALISE PROGRESS BAR
            print("\nDumping first measurement. Please wait.\n")

            # DUMP FIRST MEASUREMENT. ALWAYS GIVES UNVALID RESULT EVEN WHEN PROBE CALIBRATED.
            # rb.switchon(switch1)
            # rb.switchoff(switch2)

            # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
            # time.sleep(0.04)

            analyzer.write("SENS1:SWE:MODE SINGLE")
            analyzer.write("TRIGger:SCOPe CURRent")
            analyzer.write("INITiate1:IMMediate")
            time.sleep(OFFdelay)

            # PLACE SWITCH IN POSITION II (50 Ohm terminator)
            # rb.switchon(switch2)
            # rb.switchoff(switch1)
            time.sleep(ONdelay)

            # TURN PERISTALTIC PUMP ON
            if Peris_ON == 1:
                window.find_element('_NORMAL9_').update(text_color='red')
                voltage = RPM / 40
                # CH0 = SPEED CONTROL
                values[0].setVoltage(voltage)
                # TURN PERISTALTIC PUMP ON
                values[1].setVoltage(0.00)
                ao4.setIoGroup(channels, values)
                print("Peristaltic pump turned ON\n")

            # Turn isocratic pump ON for remainder of test (if Iso_ON = 1)
            if Iso_ON == 1:
                window.find_element('_NORMAL8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON.exe")


        else:

            print("\nITERATION " + str(i) + "\n")

            # ------------- STATE I - DIELECTRIC MEASUREMENT -------------#
            # Update next step text color on GUI window
            window.find_element('_NORMAL3_').update(text_color='black')
            window.find_element('_NORMAL5_').update(text_color='red')
            window.find_element('_NORMAL6_').update(text_color='black')
            sg.OneLineProgressMeter('Test progress...', 2 * i + 1, 2 * num_its + 4, key='METER1', grab_anywhere=True)

            """
            CHECK IF REFLECTED POWER VALUE READ IS GREATER THEN 15% OF TRANSMITTED POWER.            
            """
            rr = client.read_holding_registers(103, 1, unit=UNIT)
            rr_rpower = str(rr.registers)

            # For debugging purposes
            # print(rr_rpower[1:2])

            if (int(rr_rpower[1]) > auto_rpower and rpower == 0) or (int(rr_rpower[1:2]) > auto_rpower and rpower == 0):
                sg.popup(
                    "The reflected power is too high.\n The code will shutdown automatically.\n Rerun the code if you desire retrying the test.")
                # TURN MICROWAVES OFF
                client.write_register(2, 0x00, unit=UNIT)

                # PLACE SWITCH IN POSITION I
                # rb.switchon(switch1)
                # rb.switchoff(switch2)

                # TURN PERISTALTIC MOTOR OFF
                if Peris_ON == 1:
                    window.find_element('_NORMAL9_').update(text_color='black')
                    values[0].setVoltage(0.00)
                    values[1].setVoltage(5.00)
                    print("Peristaltic pump turned OFF\n")
                    ao4.setIoGroup(channels, values)

                # TURN ISOCRATIC PUMP OFF
                if Iso_ON == 1:
                    os.system("C:/Windows\Licop\LicopDemo\PumpOFF.exe")

                sys.exit()

            # MICROWAVES OFF
            client.write_register(2, 0x00, unit=UNIT)
            print("Microwaves OFF")

            # PLACE SWITCH IN POSITION I (Dielectric measurement)
            # rb.switchon(switch1)
            # rb.switchoff(switch2)
            print("Switch in position I (Dielectric measurements)")

            # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
            # time.sleep(0.04)

            # TURN PERISTALTIC PUMP OFF
            # if Peris_ON == 1:
            # window.find_element('_NORMAL9_').update(text_color='red')
            # values[1].setVoltage(5.00)
            # ao4.setIoGroup(channels, values)
            # print("Peristaltic pump turned OFF\n")

            # PUT ISOCRATIC PUMP ON STANDBY
            # if Iso_ON == 1:
            # os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY1.exe")

            # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .CSV (REAL IMAGINARY DATA FORMAT)
            analyzer.write("SENS1:SWE:MODE SINGLE")
            analyzer.write("TRIGger:SCOPe CURRent")
            analyzer.write("INITiate1:IMMediate")
            # analyzer.write("MMEMory:STORe:DATA '" + filename + "/Iteration" + str(i) + ".csv', 'CSV Formatted Data','Trace','RI', 1")

            # WAIT WHILE MEASUREMENTS ARE BEING TAKEN
            time.sleep(OFFdelay)

            # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .S1P (REAL IMAGINARY DATA FORMAT)
            analyzer.write("CALC:MEAS:DATA:SNP:PORTs:Save '1,,', '" + filename + "/Iteration" + str(i) + ".s1p'")
            analyzer.write("*OPC?")

            # ------------- STATE II - MICROWAVE EMISSION -------------#
            # Update next step text color on GUI window
            window.find_element('_NORMAL5_').update(text_color='black')
            window.find_element('_NORMAL6_').update(text_color='red')
            sg.OneLineProgressMeter('Test progress...', 2 * i + 2, 2 * num_its + 4, key='METER1', grab_anywhere=True)

            # PLACE SWITCH IN POSITION II (50 Ohm terminator)
            # rb.switchon(switch2)
            # rb.switchoff(switch1)
            print("Switch in position II (Microwave ablation)")

            # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ablation probe (see switch datasheet for more details)
            # time.sleep(0.04)

            # MICROWAVES ON
            client.write_register(2, 0x50, unit=UNIT)
            print("Microwaves ON")

            # TURN PERISTALTIC PUMP ON
            # if Peris_ON == 1:
            # voltage = RPM / 40
            # CH0 = SPEED CONTROL
            # values[0].setVoltage(voltage)
            # TURN PERISTALTIC PUMP ON
            # values[1].setVoltage(0.00)

            # ao4.setIoGroup(channels, values)
            # print("Peristaltic pump turned ON\n")
            # window.find_element('_NORMAL9_').update(text_color='black')

            # PUT ISOCRATIC PUMP IN ON STATE
            # if Iso_ON == 1:
            # os.system("C:/Windows\Licop\LicopDemo\PumpON.exe")

            # WAIT WHILE GENERATOR IS EMITTING MICROWAVES
            time.sleep(ONdelay)

    # -------- END OF NORMAL TEST - END STATES ---------- #
    # Update next step text color on GUI window
    window.find_element('_NORMAL5_').update(text_color='black')
    window.find_element('_NORMAL6_').update(text_color='black')
    window.find_element('_NORMAL7_').update(text_color='red')
    window.find_element('_NORMAL8_').update(text_color='black')
    window.find_element('_NORMAL9_').update(text_color='black')
    sg.OneLineProgressMeter('Test progress...', 2 * num_its + 3, 2 * num_its + 4, key='METER1', grab_anywhere=True)

    # TURN MICROWAVES OFF FOR ITS DEFAULT END STATE
    client.write_register(2, 0x00, unit=UNIT)

    # PLACE SWITCH IN POSITION I FOR ITS DEFAULT END STATE
    # rb.switchon(switch1)
    # rb.switchoff(switch2)

    # TURN PERISTALTIC MOTOR OFF FOR ITS DEFAULT END STATE
    if Peris_ON == 1:
        values[0].setVoltage(0.00)
        values[1].setVoltage(5.00)
        print("Peristaltic pump turned OFF\n")
        ao4.setIoGroup(channels, values)

    # TURN ISOCRATIC PUMP OFF FOR ITS DEFAULT END STATE
    if Iso_ON == 1:
        os.system("C:/Windows\Licop\LicopDemo\PumpOFF.exe")

    # Print end of test message
    print(
        "\n------- The test is complete! Select another test or press QUIT from the drop down menu to exit. ------- \n\n")

    client.close()

    # Update next step text color on GUI window
    time.sleep(2)
    window.find_element('_NORMAL7_').update(text_color='black')
    sg.OneLineProgressMeter('Test progress...', 2 * num_its + 4, 2 * num_its + 4, key='METER1', grab_anywhere=True)


"""
FREQUENCY SCAN SWEEP FUNCTION
"""
# AUTOMATIC FREQUENCY SWEEP MODE. SEE GENERATOR USER MANUAL FOR MORE INFO. PAGE 15. CABINET 17/18.
# TODO Fix problem with EndOfScan bit for automatic mode
def Freq_sweep(datapoints, BW, directory, powersweep, rpower, ONdelay, OFFdelay, startfreq, stopfreq,
               stepfreq, IsManu, Peris_ON, RPM, Iso_ON, IsLog):
    # MANUAL MODE
    if IsManu:
        print("------------------MANUAL SWEEP TEST------------------\n")
        time.sleep(2)
        # Initial delay prior to start of test
        print("The test will begin in 3 seconds\n")
        time.sleep(3)

        count = 0

        num_its = int(((stopfreq - startfreq) / stepfreq) + 1)

        startFreqAna = startfreq
        stopFreqAna = stopfreq

        # Update next step text color on GUI window
        window.find_element('_SWEEP2_').update(text_color='red')
        sg.OneLineProgressMeter('Test progress...', 1, 2 * num_its + 4, key='METER1', grab_anywhere=True)

        """
        1. NETWORK ANALYZER INIT
        """
        # CONVERSION DE FRÉQUENCE DE MHZ À HZ
        startFreqAnaHz = startFreqAna * 1000000
        stopFreqAnaHz = stopFreqAna * 1000000
        # Set channel 1 freq and numpoint
        analyzer.write("SENS1:SWE:POIN " + str(datapoints))
        analyzer.write("SENS1:FREQ:START " + str(startFreqAnaHz))
        analyzer.write("SENS1:FREQ:STOP " + str(stopFreqAnaHz))
        analyzer.write("SENS1:BAND " + str(BW))

        # Determine, i.e. query, number of points in trace for ASCII transfer - query
        # Verification of data transfer between host PC and vector analyzer
        analyzer.write("SENS1:SWE:POIN?")
        numPoints = analyzer.read()
        print("Number of trace points (Channel 1)\n" + numPoints)

        # Determine, i.e. query, start and stop frequencies, i.e. stimulus begin and end points
        # Verification of data transfer between host PC and vector analyzer
        analyzer.write("SENS1:FOM:RANG:FREQ:STAR?")
        analyzer.write("SENS1:FREQ:START?")
        startFreqRead = analyzer.read()
        analyzer.write("SENS1:FREQ:STOP?")
        stopFreqRead = analyzer.read()
        print(
            "Analyzer start frequency (Channel 1) = \n" + startFreqRead + "\n" + "Stop frequency (Channel 1) = \n" + stopFreqRead)

        # SET FREQUENCY SWEEP TIME (MINIMUM DELAY FOR REAL TIME APPLICATION)
        analyzer.write("SENS1:SWE:TIME " + str(OFFdelay))

        # CHANNEL 1 PUT ON HOLD MODE FOR ITS DEFAULT TRIGGER STATE
        analyzer.write("SENS1:SWE:MODE HOLD")

        # Saved format specification (RI is selected)
        analyzer.write("MMEMory:STOR:TRAC:FORM:SNP RI")

        # Selection of sweep type (linear or logarithmic)
        if IsLog == 0:
            analyzer.write("SENS1:SWEep:TYPE LIN")
            print("Linear frequency sweep selected\n")

        if IsLog == 1:
            analyzer.write("SENS1:SWEep:TYPE LOG")
            print("Logarithmic frequency sweep selected\n")

        filename = directory

        try:
            analyzer.write("mmemory:mdirectory 'D:/" + filename + "'")
            print("D: drive folder created \n")

        except:
            print("D: folders weren't created.\n")

        """
        2. GENERATOR INIT
        """
        # TIMEOUT ( Must be greater than Power_ON_time !!!)
        client.write_register(98, 3000, unit=UNIT)

        # SET START FREQUENCY
        # (24500 meaning = 24500 x 100 KHz = 2.45 Ghz)
        # See page 31 of microwave generator user manual for more information (Cabinet 18, room 455, IARC)
        startfreqKHz = startfreq * 10
        stopfreqKHz = stopfreq * 10
        stepfreqKHz = stepfreq * 10
        client.write_register(9, int(startfreqKHz), unit=UNIT)
        rr1 = client.read_holding_registers(9, 1, unit=UNIT)
        rq1 = rr1.registers

        # REFLECTED POWER LIMITATION MODE - ON
        client.write_register(2, 0x10, unit=UNIT)
        print("Reflected power limitation mode activated \n")

        # REFLECTED POWER SET
        # Value = 15% of transmitted power OR Manual selected value (See generator user manual for more details, cabinet 18, room 455, IARC)
        auto_rpower = 0.15 * powersweep
        if int(rpower) != 0:
            client.write_register(1, int(rpower), unit=UNIT)
            rr = client.read_holding_registers(1, 1, unit=UNIT)
            rr1 = rr.registers
            print("Reflected power set to " + str(rr1) + " W \n")
        else:
            client.write_register(1, int(auto_rpower), unit=UNIT)
            rr = client.read_holding_registers(1, 1, unit=UNIT)
            rr1 = rr.registers
            print("Reflected power set to " + str(rr1) + " W \n")

        # TRANSMITTED POWER SET:
        client.write_register(0, int(powersweep), unit=UNIT)
        print("Forward power set point set\n")

        # GENERATOR OFF
        client.write_register(2, 0x00, unit=UNIT)
        print("Generator turned off for its default initial state \n\n\n")

        """
        3. Peristaltic pump set
        """
        if Peris_ON == 1:
            ao4 = LucidControlAO4('COM6')

            # Open AO4 port
            if ao4.open() == False:
                print('Error connecting to port {0} '.format(
                    ao4.portName))
                ao4.close()
                exit()

            # Create a tuple of 4 voltage objects
            values = (ValueVOS4(), ValueVOS4(), ValueVOS4(), ValueVOS4())

            # Initialize a boolean tuple for channels to change.
            channels = (True, True, True, True)

        """
        4. Isocratic pump set
        """
        if Iso_ON == 1:
            sg.popup(
                "Please wait 40 seconds for isocratic pump to initialise. Once done, enter desired flow speed in command prompt. Press Okay to continue")
            os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY.exe")
            # window.findElement('_SWEEP_FRAME_').update(Visible=True)

        currentfreqKHz = startfreqKHz
        print("\n-------------START TEST-------------\n")
        print("Start frequency set to" + str(rq1) + "KHz\n")

        # ------------- DUMP FIRST MEASUREMENT - SWITCH INIT -------------#
        # Update next step text color on GUI window
        window.find_element('_SWEEP2_').update(text_color='black')
        window.find_element('_SWEEP3_').update(text_color='red')
        sg.OneLineProgressMeter('Test progress...', 2, 2 * num_its + 4, key='METER1', grab_anywhere=True)

        # INITIALISE PROGRESS BAR
        print("\nDumping first measurement. Please wait.\n")

        # DUMP FIRST MEASUREMENT. ALWAYS GIVES UNVALID RESULT EVEN WHEN PROBE CALIBRATED.
        # rb.switchon(switch1)
        # rb.switchoff(switch2)

        # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
        # time.sleep(0.04)

        analyzer.write("SENS1:SWE:MODE SINGLE")
        analyzer.write("TRIGger:SCOPe CURRent")
        analyzer.write("INITiate1:IMMediate")
        time.sleep(OFFdelay)

        # PLACE SWITCH IN POSITION II (50 Ohm terminator)
        # rb.switchon(switch2)
        # .switchoff(switch1)
        time.sleep(ONdelay)

        # TURN PERISTALTIC PUMP ON
        if Peris_ON == 1:
            window.find_element('_SWEEP9_').update(text_color='red')
            voltage = RPM / 40
            # CH0 = SPEED CONTROL
            values[0].setVoltage(voltage)
            # TURN PERISTALTIC PUMP ON
            values[1].setVoltage(0.00)
            ao4.setIoGroup(channels, values)
            print("Peristaltic pump turned ON\n")

        # Turn isocratic pump ON (if Iso_ON = 1)
        if Iso_ON == 1:
            window.find_element('_SWEEP8_').update(text_color='red')
            os.system("C:/Windows\Licop\LicopDemo\PumpON.exe")

        while currentfreqKHz <= stopfreqKHz:

            # ------------- STATE II - MICROWAVE EMISSION -------------#
            # Update next step text color on GUI window
            window.find_element('_SWEEP3_').update(text_color='black')
            window.find_element('_SWEEP5_').update(text_color='red')
            window.find_element('_SWEEP6_').update(text_color='black')
            sg.OneLineProgressMeter('Test progress...', 2 * (count + 1) + 1, 2 * num_its + 4, key='METER1',
                                    grab_anywhere=True)

            # Place switch in position II (A-D)
            # rb.switchon(switch2)
            # rb.switchoff(switch1)

            # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
            # time.sleep(0.04)

            # MICROWAVES ON
            client.write_register(2, 0x50, unit=UNIT)
            print("Microwaves ON for " + str(ONdelay) + " seconds")

            # TURN PERISTALTIC PUMP ON
            # if Peris_ON == 1:
            # window.find_element('_SWEEP9_').update(text_color='black')
            # voltage = RPM / 40
            # CH0 = SPEED CONTROL
            # values[0].setVoltage(voltage)
            # TURN PERISTALTIC PUMP ON
            # values[1].setVoltage(0.00)
            # ao4.setIoGroup(channels, values)
            # print("Peristaltic pump turned ON\n")

            # PUT ISOCRATIC PUMP IN ON STATE
            # if Iso_ON == 1:
            # os.system("C:/Windows\Licop\LicopDemo\PumpON.exe")

            # wait micro_time seconds
            time.sleep(ONdelay)

            # ------------- STATE I - DIELECTRIC MEASUREMENT -------------#
            # Update next step text color on GUI
            window.find_element('_SWEEP5_').update(text_color='black')
            window.find_element('_SWEEP6_').update(text_color='red')
            sg.OneLineProgressMeter('Test progress...', 2 * (count + 1) + 2, 2 * num_its + 4, key='METER1',
                                    grab_anywhere=True)

            # CHECK IF REFLECTED POWER VALUE READ IS GREATER THEN 30% OF TRANSMITTED POWER.
            rr = client.read_holding_registers(103, 1, unit=UNIT)
            rr_rpower = str(rr.registers)
            if (int(rr_rpower[1]) > auto_rpower and rpower == 0) or (int(rr_rpower[1:2]) > auto_rpower and rpower == 0):
                sg.popup(
                    "The reflected power is too high.\n The code will shutdown automatically.\n Rerun the code if you desire retrying the test.")
                # TURN MICROWAVES OFF
                client.write_register(2, 0x00, unit=UNIT)

                # PLACE SWITCH IN POSITION I
                # rb.switchon(switch1)
                # rb.switchoff(switch2)

                # TURN PERISTALTIC MOTOR OFF
                if Peris_ON == 1:
                    values[0].setVoltage(0.00)
                    values[1].setVoltage(5.00)
                    print("Peristaltic pump turned OFF\n")
                    ao4.setIoGroup(channels, values)

                # TURN ISOCRATIC PUMP OFF
                if Iso_ON == 1:
                    os.system("C:/Windows\Licop\LicopDemo\PumpOFF.exe")
                sys.exit()

            # TURN MICROWAVE OFF
            # Place switch in position I (A-B)
            client.write_register(2, 0x00, unit=UNIT)
            print("Microwave turned off for " + str(OFFdelay) + " seconds")

            # rb.switchon(switch1)
            # rb.switchoff(switch2)

            # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
            # time.sleep(0.04)

            # TURN PERISTALTIC PUMP OFF
            # if Peris_ON == 1:
            # window.find_element('_SWEEP9_').update(text_color='red')
            # values[1].setVoltage(5.00)
            # ao4.setIoGroup(channels, values)
            # print("Peristaltic pump turned OFF\n")

            # PUT ISOCRATIC PUMP ON STANDBY
            # if Iso_ON == 1:
            # os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY1.exe")

            # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .CSV (REAL IMAGINARY DATA FORMAT)
            analyzer.write("SENS1:SWE:MODE SINGLE")
            analyzer.write("TRIGger:SCOPe CURRent")
            analyzer.write("INITiate1:IMMediate")
            # analyzer.write("MMEMory:STORe:DATA '" + filename + "/Iteration" + str(count + 1) + ".csv', 'CSV formatted Data','Trace','RI', 1")

            # Microwave off for input value seconds (while ENA is taking measurements)
            time.sleep(OFFdelay)

            # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .S1P (REAL IMAGINARY DATA FORMAT)
            analyzer.write(
                "CALC:MEAS:DATA:SNP:PORTs:Save '1,,', '" + filename + "/Iteration" + str(count + 1) + ".s1p'")
            analyzer.write("*OPC?")

            # ------------- INCREMENT FREQUENCY -------------#

            currentfreqKHz += stepfreqKHz
            if currentfreqKHz == 2.5E9 or currentfreqKHz > stopfreqKHz:
                break
            else:
                client.write_register(9, int(currentfreqKHz), unit=UNIT)
                rr1 = client.read_holding_registers(112, 1, unit=UNIT)
                rq1 = rr1.registers
                print("New generator frequency set to :" + str(rq1) + "KHz ")
                count += 1

        # -------- END OF SWEEP MANUAL TEST - END STATES ---------- #
        # Update next step text color on GUI window
        window.find_element('_SWEEP5_').update(text_color='black')
        window.find_element('_SWEEP6_').update(text_color='black')
        window.find_element('_SWEEP7_').update(text_color='red')
        window.find_element('_SWEEP8_').update(text_color='black')
        window.find_element('_SWEEP9_').update(text_color='black')
        sg.OneLineProgressMeter('Test progress...', 2 * num_its + 3, 2 * num_its + 4, key='METER1', grab_anywhere=True)

        # TURN MICROWAVES OFF FOR ITS DEFAULT END STATE
        client.write_register(2, 0x00, unit=UNIT)

        # PLACE SWITCH IN POSITION I FOR ITS DEFAULT END STATE
        # rb.switchon(switch1)
        # rb.switchoff(switch2)

        # TURN PERISTALTIC MOTOR OFF FOR ITS DEFAULT END STATE
        if Peris_ON == 1:
            values[0].setVoltage(0.00)
            values[1].setVoltage(5.00)
            ao4.setIoGroup(channels, values)

        # TURN ISOCRATIC PUMP OFF FOR ITS DEFAULT END STATE
        if Iso_ON == 1:
            os.system("C:/Windows\Licop\LicopDemo\PumpOFF.exe")

        # Update next step text color on GUI window
        time.sleep(2)
        window.find_element('_SWEEP7_').update(text_color='black')
        sg.OneLineProgressMeter('Test progress...', 2 * num_its + 4, 2 * num_its + 4, key='METER1', grab_anywhere=True)

        # Print end of test message
        print(
            "\n------- The test is complete! Select another test or press QUIT from the drop down menu to exit. ------- \n\n")

    # AUTOMATIC MODE
    if not IsManu:
        print("------------------AUTOMATIC SWEEP TEST------------------\n")
        time.sleep(2)
        # Initial delay prior to start of test
        print("The test will begin in 3 seconds\n")
        time.sleep(3)

        # TIMEOUT ( Must be greater than Power_ON_time !!!)
        client.write_register(98, 3000, unit=UNIT)

        startfreqKHz = startfreq * 10
        stopfreqKHz = stopfreq * 10
        stepfreqKHz = stepfreq * 10

        # SET START, STOP and FREQ STEP in 100 KHz format
        client.write_register(11, int(startfreqKHz), unit=UNIT)
        rr1 = client.read_holding_registers(11, 1, unit=UNIT)
        rq1 = rr1.registers
        print("Start frequency set to:\n " + str(rq1) + "KHz\n")

        client.write_register(12, int(stopfreqKHz), unit=UNIT)
        rr2 = client.read_holding_registers(12, 1, unit=UNIT)
        rq2 = rr2.registers
        print("Stop frequency set to:\n " + str(rq2) + "KHz\n")

        client.write_register(16, int(stepfreqKHz), unit=UNIT)
        rr3 = client.read_holding_registers(16, 1, unit=UNIT)
        rq3 = rr3.registers
        print("Frequency step set to:\n" + str(rq3) + "MHz\n")

        # ACTIVATE SCAN MODE - SET SCAN MODE BIT TO 1
        client.write_register(17, 1, unit=UNIT)
        # rr = client.read_holding_registers(17, 1, unit=UNIT)
        # rq = rr.registers
        # print(rq)

        # REFLECTED POWER LIMITATION MODE - ON
        client.write_register(2, 0x10, unit=UNIT)

        # REFLECTED POWER SET
        client.write_register(1, 20, unit=UNIT)

        # TRANSMITTED POWER SET
        client.write_register(0, int(powersweep), unit=UNIT)

        # GENERATOR ON
        # Switch position II
        # rb.switchon(switch2)
        # rb.switchoff(switch1)

        client.write_register(2, 0x50, unit=UNIT)
        print("Generator turned ON\n")
        """
        """

        while True:
            rr = client.read_holding_registers(105, 1, unit=UNIT)
            rq = rr.registers
            print("End of scan response is: " + str(rq))
            time.sleep(2)

            """
            rr = client.read_holding_registers(113, 1, unit=UNIT)
            rq = rr.registers
            print("Currrent scan frequency is:" + str(rq))         
            """

            if rq[0] == 160 or rq[0] == 32 or rq[0] == 224:
                print("The scan is complete! Processing data\n")
                break

        # ACTIVATE SCAN DATA MODE - SET SCAN DATA BIT TO 1
        client.write_register(17, 0x04, unit=UNIT)

        while True:
            rr = client.read_holding_registers(109, 1, unit=UNIT)
            rq = rr.registers
            print(rq)
            time.sleep(5)
            if rq[0] == 0:
                rr = client.read_holding_registers(114, 1, unit=UNIT)
                rq = rr.registers
                print("Scanned frequency minimum 1 data:\n " + str(rq) + "\n")

                rr = client.read_holding_registers(115, 1, unit=UNIT)
                rq = rr.registers
                print("Scanned frequency minimum 1 data:\n " + str(rq) + "\n")

                rr = client.read_holding_registers(113, 1, unit=UNIT)
                rq = rr.registers
                print("Scanned frequency register:\n " + str(rq) + "\n")
                time.sleep(0.1)
                break

        # Turn generator OFF for its default end state
        # Place switch in position I (A-B) for its default end state
        client.write_register(2, 0x00, unit=UNIT)
        # rb.switchon(switch1)
        # rb.switchoff(switch2)

        client.close()
        print("The Sairem automatic frequency sweep test is complete!\n\n")


"""
MANUAL GENERATOR ON FUNCTION
"""
# GENERATOR + SWITCH ONLY
def Manual_Gen_ON(power_man, rpower):
    # Place switch in position II (A-D)
    # rb.switchon(switch2)
    # rb.switchoff(switch1)

    # TIMEOUT (30 seconds)
    client.write_register(98, 3000, unit=UNIT)

    # SET FREQUENCY
    freq = 24500
    client.write_register(9, freq, unit=UNIT)

    # REFLECTED POWER LIMITATION MODE - ON
    client.write_register(2, 0x10, unit=UNIT)

    # REFLECTED POWER SET
    # Value = 30% of transmitted power OR Manual selected value (See generator user manual for more details, cabinet 18, room 455, IARC)
    auto_rpower = 0.25 * power_man
    if int(rpower) != 0:
        client.write_register(1, int(rpower), unit=UNIT)
        rr = client.read_holding_registers(1, 1, unit=UNIT)
        rr1 = rr.registers
        print("Reflected power set to " + str(rr1) + " W \n")
    else:
        client.write_register(1, int(auto_rpower), unit=UNIT)
        rr = client.read_holding_registers(1, 1, unit=UNIT)
        rr1 = rr.registers
        print("Reflected power set to " + str(rr1) + " W \n")

    # TRANSMITTED POWER SET
    client.write_register(0, power_man, unit=UNIT)

    # GENERATOR ON
    client.write_register(2, 0x50, unit=UNIT)
    print("Generator turned ON\n")

    # Update dynamic text display
    window.find_element('OFFSETTEXT').update(text_color='black')
    window.find_element('ONSETTEXT').update(text_color='red')


"""
MANUAL GENERATOR OFF FUNCTION
"""
# GENERATOR + SWTICH ONLY
def Manual_Gen_OFF():
    # GENERATOR OFF
    client.write_register(2, 0x00, unit=UNIT)
    print("Generator turned OFF\n")

    # Update dynamic text display
    window.find_element('OFFSETTEXT').update(text_color='red')
    window.find_element('ONSETTEXT').update(text_color='black')
    # Place switch in position I (A-B)
    # rb.switchon(switch1)
    # rb.switchoff(switch2)


"""
MANUAL ENA TRIGGER FUNCTION
"""
# ENA + SWITCH ONLY
def Manual_ENA_Trigger(startFreq, stopFreq, datapoints, BW, delay, directory, IsLog):
    # CONVERSION DE FRÉQUENCE DE MHZ À HZ
    startFreqHz = startFreq * 1000000
    stopFreqHz = stopFreq * 1000000
    # Set channel 1 freq and numpoint
    analyzer.write("SENS1:SWE:POIN " + str(datapoints))
    analyzer.write("SENS1:FREQ:START " + str(startFreqHz))
    analyzer.write("SENS1:FREQ:STOP " + str(stopFreqHz))
    analyzer.write("SENS1:BAND " + str(BW))

    # Determine, i.e. query, number of points in trace for ASCII transfer - query
    # Verification of data transfer between host PC and vector analyzer
    analyzer.write("SENS1:SWE:POIN?")
    numPoints = analyzer.read()
    print("Number of trace points (Channel 1)\n" + numPoints)

    # Determine, i.e. query, start and stop frequencies, i.e. stimulus begin and end points
    # Verification of data transfer between host PC and vector analyzer
    analyzer.write("SENS1:FOM:RANG:FREQ:STAR?")
    analyzer.write("SENS1:FREQ:START?")
    startFreq = analyzer.read()
    analyzer.write("SENS1:FREQ:STOP?")
    stopFreq = analyzer.read()
    print("Analyzer start frequency (Channel 1) = \n" + startFreq + "\n" + "Stop frequency (Channel 1) = \n" + stopFreq)

    # SET FREQUENCY SWEEP TIME (MINIMUM DELAY FOR REAL TIME APPLICATION)
    analyzer.write("SENS1:SWE:TIME " + str(delay))

    # CHANNEL 1 PUT ON HOLD MODE FOR ITS DEFAULT TRIGGER STATE
    analyzer.write("SENS1:SWE:MODE HOLD")

    # Saved format specification (RI is selected)
    analyzer.write("MMEMory:STOR:TRAC:FORM:SNP RI")

    filename = directory

    # Selection of sweep type (linear or logarithmic)
    if IsLog == 0:
        analyzer.write("SENS1:SWEep:TYPE LIN")
        print("Linear frequency sweep selected\n")

    if IsLog == 1:
        analyzer.write("SENS1:SWEep:TYPE LOG")
        print("Logarithmic frequency sweep selected\n")

    try:
        analyzer.write("mmemory:mdirectory 'D:/" + filename + "'")
        print("D: drive folder created \n")

    except:
        print("D: folders weren't created.\n")

    # Place switch in position I (A-B)
    # rb.switchon(switch1)
    # rb.switchoff(switch2)

    # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from dielectric probe (see switch datasheet for more details)
    # time.sleep(0.04)

    # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .CSV (REAL IMAGINARY DATA FORMAT)
    analyzer.write("SENS1:SWE:MODE SINGLE")
    analyzer.write("TRIGger:SCOPe CURRent")
    analyzer.write("INITiate1:IMMediate")
    # analyzer.write("MMEMory:STORe:DATA '" + filename + "/ENAManTrig_Result" + ".csv', 'CSV formatted Data','Trace','RI', 1")

    time.sleep(delay)
    # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .S1P (REAL IMAGINARY DATA FORMAT)
    analyzer.write("CALC:MEAS:DATA:SNP:PORTs:Save '1,,', '" + filename + "/ENAManTrig_Result" + ".s1p'")
    analyzer.write("*OPC?")


"""
MANUAL PERISTALTIC AND ISOCRATIC PUMP CONTROL
"""


# PUMPS ONLY (NO SWITCH CONFIGURATION NEEDED)
def Manual_Pumps(Manu_Pumps_State, RPM, Manu_Iso_ON, Manu_Peris_ON, voltageOutput0, voltageOutput1):

    voltagePump = RPM / 40

    # ISOCRATIC PUMP ON
    if Manu_Pumps_State == 1:
        # sg.popup("Please wait 40 seconds for isocratic pump to initialise. Once done, enter desired flow speed in command prompt. Press Okay to continue")
        # os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY.exe")
        os.system("C:/Windows\Licop\LicopDemo\PumpON.exe")
        if Manu_Peris_ON == 1:
            window.find_element('_ALCOHOLON_').update(text_color='red')
            window.find_element('_ALCOHOLOFF_').update(text_color='black')
            window.find_element('_COOLINGON_').update(text_color='red')
            window.find_element('_COOLINGOFF_').update(text_color='black')
        elif Manu_Peris_ON == 0:
            window.find_element('_ALCOHOLON_').update(text_color='red')
            window.find_element('_ALCOHOLOFF_').update(text_color='black')
            window.find_element('_COOLINGON_').update(text_color='black')
            window.find_element('_COOLINGOFF_').update(text_color='red')


    # PERISTALTIC PUMP ON
    if Manu_Pumps_State == 3:

        if Manu_Iso_ON == 1 :
            window.find_element('_ALCOHOLON_').update(text_color='red')
            window.find_element('_ALCOHOLOFF_').update(text_color='black')
            window.find_element('_COOLINGON_').update(text_color='red')
            window.find_element('_COOLINGOFF_').update(text_color='black')
        elif Manu_Iso_ON == 0:
            window.find_element('_ALCOHOLON_').update(text_color='black')
            window.find_element('_ALCOHOLOFF_').update(text_color='red')
            window.find_element('_COOLINGON_').update(text_color='red')
            window.find_element('_COOLINGOFF_').update(text_color='black')

        voltageOutput0.setDeviceSerialNumber(589734)
        voltageOutput0.setChannel(0)
        voltageOutput1.setDeviceSerialNumber(589734)
        voltageOutput1.setChannel(1)

        voltageOutput0.openWaitForAttachment(1000)
        voltageOutput1.openWaitForAttachment(1000)
        voltageOutput0.setVoltage(0)
        voltageOutput1.setVoltage(voltagePump)


    # ISOCRATIC PUMP OFF
    if Manu_Pumps_State == 2:
        os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY.exe")
        if Manu_Peris_ON == 1:
            window.find_element('_ALCOHOLON_').update(text_color='black')
            window.find_element('_ALCOHOLOFF_').update(text_color='red')
            window.find_element('_COOLINGON_').update(text_color='red')
            window.find_element('_COOLINGOFF_').update(text_color='black')
        elif Manu_Peris_ON == 0:
            window.find_element('_ALCOHOLON_').update(text_color='black')
            window.find_element('_ALCOHOLOFF_').update(text_color='red')
            window.find_element('_COOLINGON_').update(text_color='black')
            window.find_element('_COOLINGOFF_').update(text_color='red')


    # PERISTALTIC PUMP OFF
    if Manu_Pumps_State == 4:

        if Manu_Iso_ON == 1:
            window.find_element('_ALCOHOLON_').update(text_color='red')
            window.find_element('_ALCOHOLOFF_').update(text_color='black')
            window.find_element('_COOLINGON_').update(text_color='black')
            window.find_element('_COOLINGOFF_').update(text_color='red')
        elif Manu_Iso_ON == 0:
            window.find_element('_ALCOHOLON_').update(text_color='black')
            window.find_element('_ALCOHOLOFF_').update(text_color='red')
            window.find_element('_COOLINGON_').update(text_color='black')
            window.find_element('_COOLINGOFF_').update(text_color='red')


        voltageOutput0.setDeviceSerialNumber(589734)
        voltageOutput0.setChannel(0)
        voltageOutput1.setDeviceSerialNumber(589734)
        voltageOutput1.setChannel(1)

        voltageOutput0.openWaitForAttachment(1000)
        voltageOutput1.openWaitForAttachment(1000)

        # SET RPM SPEED AND TURN OFF PUMP
        # CH1 = SPEED CONTROL
        # TURN PERISTALTIC PUMP OFF
        voltageOutput0.setVoltage(5)
        voltageOutput1.setVoltage(0)


"""
RESET GENERATOR FAULTS FUNCTION
"""
def Reset_Faults():
    # RESET FAULT : set bit 7 of register 2 to 1 and then to 0
    # (see page 19 of Microwave generator documentation --> Cabinet 18, room 455, IARC)
    client.write_register(2, 0x80, unit=UNIT)

    # RESET FAULT: set bit 7 of register 2 back to 0
    # TURN GENERATOR OFF (default state) : set register 2, bit 6 to
    client.write_register(2, 0x00, unit=UNIT)
    print("\nFaults Reseted\n")
    sg.popup("Fault reseted")


"""
FIVE ITERATION TEST FUNCTION
"""
def Five_Iteration_Test(startFreq, stopFreq, datapoints, BW, directory, delaimicro, delaimesure, voltageOutput0, voltageOutput1, power, ONdelay, Flow,
                        freq, rpower, i, Peris_ON, RPM, Iso_ON, IsLog, LogFileName, DonneesTemps, rpowergraph, powergraph, Dielec_Verif, min_dielec_value,
                        Timer_ON):

    filename = directory

    num_its = 5

    auto_rpower = 0.5 * power

    if i == 1:

        """
        Peristaltic pump set

        Pompe s'allume aussitôt que le code est lancé. Proposition suggéré par Jocelyn et Gabriel.
        """
        if Peris_ON == 1:

            voltageOutput0.setDeviceSerialNumber(589734)
            voltageOutput0.setChannel(0)
            voltageOutput1.setDeviceSerialNumber(589734)
            voltageOutput1.setChannel(1)

            voltageOutput0.openWaitForAttachment(1000)
            voltageOutput1.openWaitForAttachment(1000)

            
            # SET RPM SPEED AND TURN ON PUMP
            # CH1 = SPEED CONTROL CONFIG
            # CH0 = TURN PERISTALTIC PUMP ON
            voltagePump = RPM / 40
            voltageOutput0.setVoltage(0)
            voltageOutput1.setVoltage(voltagePump)

            window.find_element('_FIVEIT9_').update(text_color='red')
            # TextFile.write("\nPeristaltic pump turned ON\n")
            print("Peristaltic pump turned ON\n")


        # Créer et ouvrir fichier text pour sauvegarder le log des opérations
        TextFile = open("D:\Ablation_Automatisation\Programmation\Automatisation_Andre\Logs/" + str(LogFileName) + ".txt", "w")

        print("------------------FIVE ITERATION TEST------------------\n")
        TextFile.write("------------------FIVE ITERATION TEST------------------\n")
        time.sleep(2)

        TextFile.write("\nToday's date and time is :\n" + str(LogFileName) + "\n")

        # INITIAL DELAY
        print("\nThe test will begin in 3 seconds\n")
        TextFile.write("\nThe test will begin in 3 seconds\n\n")
        time.sleep(3)

        print("Initialising system\n")
        TextFile.write("Initialising system\n")

        # Update next step text color on GUI window
        window.find_element('_FIVEIT2_').update(text_color='red')
        sg.OneLineProgressMeter('Test progress...', 1, num_its + 4, key='METER1', grab_anywhere=True)

        """
        1. NETWORK ANALYZER SET
        """
        # CONVERSION DE FRÉQUENCE DE MHZ À HZ
        startFreqHz = startFreq * 1000000
        stopFreqHz = stopFreq * 1000000
        # Set channel 1 freq and numpoint
        analyzer.write("SENS1:SWE:POIN " + str(datapoints))
        analyzer.write("SENS1:FREQ:START " + str(startFreqHz))
        analyzer.write("SENS1:FREQ:STOP " + str(stopFreqHz))
        analyzer.write("SENS1:BAND " + str(BW))

        # Determine, i.e. query, number of points in trace for ASCII transfer - query
        # Verification of data transfer between host PC and vector analyzer
        analyzer.write("SENS1:SWE:POIN?")
        numPoints = analyzer.read()

        print("Number of trace points (Channel 1)\n" + numPoints)
        TextFile.write("\n\n----------NETWORK ANALYZER INIT------------\n\n")
        TextFile.write("Number of trace points (Channel 1)\n" + str(numPoints) + "\n")


        # Determine, i.e. query, start and stop frequencies, i.e. stimulus begin and end points
        # Verification of data transfer between host PC and vector analyzer
        analyzer.write("SENS1:FOM:RANG:FREQ:STAR?")
        analyzer.write("SENS1:FREQ:START?")
        startFreq = analyzer.read()
        analyzer.write("SENS1:FREQ:STOP?")
        stopFreq = analyzer.read()

        print("Analyzer start frequency (Channel 1) = \n" + startFreq + "\n" + "Stop frequency (Channel 1) = \n" + stopFreq)
        TextFile.write("Analyzer start frequency (Channel 1) = \n" + str(startFreq) + "\n" + "Stop frequency (Channel 1) = \n" + str(stopFreq) + "\n")

        # SET FREQUENCY SWEEP TIME (MINIMUM DELAY FOR REAL TIME APPLICATION)
        analyzer.write("SENS1:SWE:TIME " + str(delaimesure))
        TextFile.write("Measurement sweep set to " + str(delaimesure) + " seconds\n\n")

        # CHANNEL 1 PUT ON HOLD MODE FOR ITS DEFAULT TRIGGER STATE
        analyzer.write("SENS1:SWE:MODE HOLD")
        TextFile.write("Sweep mode set to HOLD\n\n")

        # Saved format specification (RI is selected)
        analyzer.write("MMEMory:STOR:TRAC:FORM:SNP RI")

        # Selection of sweep type (linear or logarithmic)
        if IsLog == 0:
            analyzer.write("SENS1:SWEep:TYPE LIN")
            print("Linear frequency sweep selected\n")
            TextFile.write("Linear frequency sweep selected\n\n")

        if IsLog == 1:
            analyzer.write("SENS1:SWEep:TYPE LOG")
            print("Logarithmic frequency sweep selected\n")
            TextFile.write("Logarithmic frequency sweep selected\n\n")


        analyzer.write("mmemory:mdirectory 'D:/" + filename + "'")
        print("D: drive folder created \n")
        TextFile.write("D: drive folder created under the name " + str(filename) + "\n\n")

        """
        2. GENERATOR SET
        """
        # TIMEOUT
        client.write_register(98, 3000, unit=UNIT)
        print("Generator Timeout set to 300 seconds\n")
        TextFile.write("\n\n----------GENERATOR INIT------------\n\n")
        TextFile.write("\nGenerator Timeout set to 300 seconds\n\n")

        # FREQUENCY SET
        freq_KHz = freq * 10
        client.write_register(9, freq_KHz, unit=UNIT)
        print("Generator frequency is set to :\n" + str(freq) + " MHz\n")
        TextFile.write("Generator frequency is set to :\n" + str(freq) + " MHz\n\n")

        # REFLECTED POWER LIMITATION MODE ON
        client.write_register(2, 0x10, unit=UNIT)
        print("Reflected power limitation mode activated \n")
        TextFile.write("Reflected power limitation mode activated \n\n")

        # TURN GENERATOR OFF (default state)
        client.write_register(2, 0x00, unit=UNIT)
        print("Microwave is turned off for its initial state \n")
        TextFile.write("Microwave is turned off for its initial state \n")

        """
        3. Isocratic pump set
        NOTE: Pu besoin d'initialiser la pompe isocratique. La pompe demeurera allumer 24/7 dans son état STANDBY.
        Proposition suggéré par Jocelyn et Gabriel.
        Si la pompe ne demeure pas dans son état STANDBY lorsque le test est lancé, la pompe ne s'allumera qu'approximativement
        30 secondes à la suite du début de son lancement.
        """
        # if Iso_ON == 1:
            # sg.popup("Please wait 40 seconds for isocratic pump to initialise. Press Okay to continue")
            # os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY.exe")


        print("\n---------------------START TEST---------------------\n")
        TextFile.write("\n\n---------------------START TEST---------------------\n\n")


        # -------------- DUMP FIRST MEASUREMENT - SWITCH INIT ----------------#

        # Update next step text color on GUI window
        window.find_element('_FIVEIT2_').update(text_color='black')
        window.find_element('_FIVEIT3_').update(text_color='red')
        sg.OneLineProgressMeter('Test progress...', 2, num_its + 4, key='METER1', grab_anywhere=True)

        print("\nDumping first measurement. Please wait.\n")
        TextFile.write("\nDumping first measurement. Please wait.\n")

        # DUMP FIRST MEASUREMENT. ALWAYS GIVES UNVALID RESULT EVEN WHEN PROBE CALIBRATED.
        # rb.switchon(switch1)
        # rb.switchoff(switch2)

        # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
        # time.sleep(0.04)

        analyzer.write("SENS1:SWE:MODE SINGLE")
        analyzer.write("TRIGger:SCOPe CURRent")
        analyzer.write("INITiate1:IMMediate")
        time.sleep(delaimesure)

        # PLACE SWITCH IN POSITION II (50 Ohm terminator)
        # rb.switchon(switch2)
        # rb.switchoff(switch1)

        # time.sleep(delaimicro)


    else:

        # Lorsque la valeur de -1 est choisie dans le drop down menu, l'étape respective de la procédure sera sauter
        # Alors, ca permet a l'opérateur de choisir le nombre d'étape à sa procédure sans à avoir besoin de modifier le code
        if Flow != -1:

            # Créer et ouvrir fichier text pour sauvegarder le log des opérations
            TextFile = open("D:\Ablation_Automatisation\Programmation\Automatisation_Andre\Logs/" + str(LogFileName) + ".txt", "a")

            """
            RUN TEST
            """

            print("\nITERATION " + str(i - 1) + "\n")
            TextFile.write("\n\nITERATION " + str(i - 1) + "\n\n")

            # Turn isocratic pump ON with flow at 1 ml/min
            if Iso_ON == 1 and Flow == 0:
                window.find_element('_FIVEIT8_').update(text_color='black')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow0.exe")

            # Turn isocratic pump ON with flow at 1 ml/min
            if Iso_ON == 1 and Flow == 1:
                window.find_element('_FIVEIT8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow1.exe")
                TextFile.write("\n Pump Turned ON with flow of 1 ml/min\n\n")

            # Turn isocratic pump ON with flow at 2 ml/min
            elif Iso_ON == 1 and Flow == 2:
                window.find_element('_FIVEIT8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow2.exe")
                TextFile.write("\n Pump Turned ON with flow of 2 ml/min\n\n")

            # Turn isocratic pump ON with flow at 1 ml/min
            if Iso_ON == 1 and Flow == 2.5:
                window.find_element('_FIVEIT8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow2.5.exe")
                TextFile.write("\n Pump Turned ON with flow of 2.5 ml/min\n\n")

            # Turn isocratic pump ON with flow at 3 ml/min
            elif Iso_ON == 1 and Flow == 3:
                window.find_element('_FIVEIT8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow3.exe")
                TextFile.write("\n Pump Turned ON with flow of 3 ml/min\n\n")


            # Turn isocratic pump ON with flow at 4 ml/min
            elif Iso_ON == 1 and Flow == 4:
                window.find_element('_FIVEIT8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow4.exe")
                TextFile.write("\n Pump Turned ON with flow of 4 ml/min\n\n")


            # Turn isocratic pump ON with flow at 5 ml/min
            elif Iso_ON == 1 and Flow == 5:
                window.find_element('_FIVEIT8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow5.exe")
                TextFile.write("\n Pump Turned ON with flow of 5 ml/min\n\n")


            # Turn isocratic pump ON with flow at 10 ml/min
            elif Iso_ON == 1 and Flow == 10:
                window.find_element('_FIVEIT8_').update(text_color='red')
                os.system("C:/Windows\Licop\LicopDemo\PumpON_Flow10.exe")
                TextFile.write("\n Pump Turned ON with flow of 10 ml/min\n\n")


            # REFLECTED POWER SET
            # Value = 50% of transmitted power OR Manual selected value (See generator user manual for more details, cabinet 18, room 455, IARC)
            if int(rpower) != 0:
                client.write_register(1, int(rpower), unit=UNIT)
                rr = client.read_holding_registers(1, 1, unit=UNIT)
                rr1 = rr.registers
                print("     Reflected power set to " + str(rr1) + " W \n")
                TextFile.write("    Reflected power set to " + str(rr1) + " W \n")

            else:
                client.write_register(1, int(auto_rpower), unit=UNIT)
                rr = client.read_holding_registers(1, 1, unit=UNIT)
                rr1 = rr.registers
                print("      Max reflected power set to " + str(rr1) + " W \n")
                TextFile.write("     Max reflected power set to " + str(rr1) + " W \n")

            # TRANSMITTED POWER SET
            client.write_register(0, int(power), unit=UNIT)
            print("      Microwave output power set to: \n     " + str(power) + " Watts\n")
            TextFile.write("     Microwave output power set to: \n     " + str(power) + " Watts\n")

            TempsTot = 0
            break_ON = 0

            while TempsTot < float(ONdelay - (delaimesure + delaimicro)):

                # Ouvrir fichier text pour sauvegarder le log des opérations entre chaque mesure
                TextFile = open("D:\Ablation_Automatisation\Programmation\Automatisation_Andre\Logs/" + str(LogFileName) + ".txt","a")

                # Ajoutée valeurs de temps en temps réel
                TempsMtn1 = str(datetime.datetime.now(pytz.timezone('America/Moncton')))
                TempsMtnFloat1 = float(TempsMtn1[17:22])
                timenow1 = TempsMtn1[11:22]
                TextFile.write("\nTime:\n" + str(timenow1) + "\n\n")
                DonneesTemps.append(timenow1)


                # ------------- STATE I - DIELECTRIC MEASUREMENT -------------#

                # Update next step text color on GUI window
                window.find_element('_FIVEIT3_').update(text_color='black')
                window.find_element('_FIVEIT5_').update(text_color='black')
                window.find_element('_FIVEIT6_').update(text_color='red')
                sg.OneLineProgressMeter('Test progress...', i + 1, num_its + 4, key='METER1', grab_anywhere=True)

                # CHECK IF REFLECTED POWER VALUE READ IS GREATER THEN 50% OF TRANSMITTED POWER WHEN RPOWER IS AUTOMATICALLY CONFIGURED
                rr = client.read_holding_registers(103, 1, unit=UNIT)
                rr_rpower = rr.registers
                if int(rr_rpower[0]) > auto_rpower and rpower == 0 :
                    # TURN MICROWAVES OFF
                    client.write_register(2, 0x00, unit=UNIT)

                    # PLACE SWITCH IN POSITION I
                    # rb.switchon(switch1)
                    # rb.switchoff(switch2)

                    # TURN PERISTALTIC MOTOR OFF
                    # if Peris_ON == 1:

                        # voltageOutput0.setDeviceSerialNumber(589734)
                        # voltageOutput0.setChannel(0)
                        # voltageOutput1.setDeviceSerialNumber(589734)
                        # voltageOutput1.setChannel(1)

                        # voltageOutput0.openWaitForAttachment(1000)
                        # voltageOutput1.openWaitForAttachment(1000)

            
                        # SET RPM SPEED AND TURN ON PUMP
                        # CH1 = SPEED CONTROL CONFIG
                        # CH0 = TURN PERISTALTIC PUMP OFF
                        # voltageOutput0.setVoltage(5)
                        # voltageOutput1.setVoltage(0)

                        # print("Peristaltic pump turned OFF\n")

                    # TURN ISOCRATIC PUMP OFF
                    if Iso_ON == 1:
                        os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY.exe")
                    
                    sg.popup("The reflected power is too high.\nThe generator has shutdown automatically.\nCooling pump will continue to flow.\nData can still be retrieved")
                    return 1

                # MICROWAVES OFF
                # Place switch in position I (AB , Dielectric measurement)
                client.write_register(2, 0x00, unit=UNIT)
                # rb.switchon(switch1)
                # rb.switchoff(switch2)

                # Delai rise and fall 
                time.sleep(0.001)

                rr = client.read_holding_registers(102, 1, unit=UNIT)
                rr2 = rr.registers

                # Sortir de la boucle si la puissance transmise actuelle n'est pas mesurer a une valeur de 0 Watts
                if rr2[0] != 0:
                    # TURN MICROWAVES OFF
                    client.write_register(2, 0x00, unit=UNIT)
                    # PLACE SWITCH IN POSITION I
                    # rb.switchon(switch1)
                    # rb.switchoff(switch2)

                    # TURN PERISTALTIC MOTOR OFF
                    # if Peris_ON == 1:

                        # voltageOutput0.setDeviceSerialNumber(589734)
                        # voltageOutput0.setChannel(0)
                        # voltageOutput1.setDeviceSerialNumber(589734)
                        # voltageOutput1.setChannel(1)

                        # voltageOutput0.openWaitForAttachment(1000)
                        # voltageOutput1.openWaitForAttachment(1000)

            
                        # SET RPM SPEED AND TURN ON PUMP
                        # CH1 = SPEED CONTROL CONFIG
                        # CH0 = TURN PERISTALTIC PUMP OFF
                        # voltageOutput0.setVoltage(5)
                        # voltageOutput1.setVoltage(0)

                        # print("Peristaltic pump turned OFF\n")

                    # TURN ISOCRATIC PUMP OFF
                    if Iso_ON == 1:
                        os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY.exe")

                    sg.popup("Power is not at 0 when it should be.\nThe generator has shutdown automatically.\nCooling pump will continue to flow.\nData can still be retrieved")

                    return 1

                # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
                # time.sleep(0.04)

                print("     Microwaves OFF for " + str(delaimesure) + " seconds")
                TextFile.write("\n      Microwaves OFF for " + str(delaimesure) + " seconds")
                # print("     Switch in position I (Dielectric measurement)\n")
                # TextFile.write("\n      Switch in position I (Dielectric measurement)\n")

                # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .CSV (REAL IMAGINARY DATA FORMAT)
                analyzer.write("SENS1:SWE:MODE SINGLE")
                analyzer.write("TRIGger:SCOPe CURRent")
                analyzer.write("INITiate1:IMMediate")
                # analyzer.write("MMEMory:STORe:DATA '" + filename + "/Iteration" + str(i - 1) + "_" + str(j) + ".csv', 'CSV formatted Data','Trace','RI', 1")

                # DELAY OF INPUT GUI VALUE SECONDS
                time.sleep(delaimesure)

                # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .S1P (REAL IMAGINARY DATA FORMAT)
                analyzer.write("CALC:MEAS:DATA:SNP:PORTs:Save '1,,', '" + filename + "/Iteration_" + str(len(DonneesTemps)) + ".s1p'")
                analyzer.write("*OPC?")
                print("     Measure triggered and saved\n")
                TextFile.write("\n      Measure triggered and saved\n")



                # ------------- STATE II - MICROWAVE ABLATION -------------#

                # Update next step text color on GUI window
                window.find_element('_FIVEIT5_').update(text_color='red')
                window.find_element('_FIVEIT6_').update(text_color='black')
                sg.OneLineProgressMeter('Test progress...', i + 1, num_its + 4, key='METER1', grab_anywhere=True)

                # MICROWAVES ON
                # Place switch in position II (AD , Microwave ablation)
                # rb.switchon(switch2)
                # rb.switchoff(switch1)

                # IMPORTANT: WAIT 40 ms for switch to stabilise electrical signal received from ENA (see switch datasheet for more details)
                # time.sleep(0.04)

                client.write_register(2, 0x50, unit=UNIT)
                # print("     Switch in position II (Microwave ablation)")
                # TextFile.write("\n      Switch in position II (Microwave ablation)\n")
                print("        Microwaves ON for " + str(delaimicro) + " seconds\n")
                TextFile.write("      Microwaves ON for " + str(delaimicro) + " seconds\n")

                # DELAY OF GUI INPUT VALUE SECONDS
                time.sleep(delaimicro)

                # Ajoutée valeurs de puissance réfléchie et transmise en temps réel à leurs listes respectivess
                rr = client.read_holding_registers(103, 1, unit=UNIT)
                rr1 = rr.registers
                rpowergraph.append(rr1)
                TextFile.write("      Reflected power measurement: " + str(rr1[0]) + " Watts\n")

                rr = client.read_holding_registers(102, 1, unit=UNIT)
                rr2 = rr.registers
                powergraph.append(rr2)
                TextFile.write("      Transmitted power measurement: " + str(rr2[0]) + " Watts\n\n")

                # VERIFIER SI LA VALEUR DIELECTRIQUE DESIREE EST ATTEINTE. Si oui, arreter test.
                if Dielec_Verif == 1:
                    break_ON = Dielec_Data_Verif(min_dielec_value, datapoints, directory)

                    if break_ON == 1:
                        return 2
                    else:
                        pass

                # PRENDRE MESURE DE TEMPS ACTUELLE AFIN DE DÉTERMINER SI BESOIN DE SORTIR DE LA BOUCLE OU NON
                TempsMtn = str(datetime.datetime.now(pytz.timezone('America/Moncton')))
                TempsMtnFloat = float(TempsMtn[17:22])

                if TempsMtnFloat - TempsMtnFloat1 < 0:
                    TempsTot += (TempsMtnFloat - TempsMtnFloat1 + 60)

                else:
                    TempsTot += (TempsMtnFloat - TempsMtnFloat1)



            # Si le temps une fois sortie de la boucle est plus petit que le temps de l'étape, attendre X secondes afin d'attendre que le délai de l'étape
            # soit complété.
            if TempsTot < ONdelay:

                # Ouvrir fichier text pour sauvegarder le log des opérations entre chaque mesure
                TextFile = open("D:\Ablation_Automatisation\Programmation\Automatisation_Andre\Logs/" + str(LogFileName) + ".txt","a")

                # Ajoutée valeurs de temps en temps réel
                TempsMtn1 = str(datetime.datetime.now(pytz.timezone('America/Moncton')))
                timenow1 = TempsMtn1[11:22]
                TextFile.write("\n\nEnd of loop time:\n" + str(timenow1) + "\n\n")
                DonneesTemps.append(timenow1)

                # Trigger final S11 value at the end of the loop
                # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .s1p (REAL IMAGINARY DATA FORMAT)
                analyzer.write("SENS1:SWE:MODE SINGLE")
                analyzer.write("TRIGger:SCOPe CURRent")
                analyzer.write("INITiate1:IMMediate")

                # DELAY OF INPUT GUI VALUE SECONDS
                time.sleep(delaimesure)

                # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .S1P (REAL IMAGINARY DATA FORMAT)
                analyzer.write("CALC:MEAS:DATA:SNP:PORTs:Save '1,,', '" + filename + "/Iteration_" + str(len(DonneesTemps)) + ".s1p'")
                analyzer.write("*OPC?")
                print("     End of loop measure triggered and saved\n")
                TextFile.write("\n      End of loop measure triggered and saved\n\n")


                rr = client.read_holding_registers(103, 1, unit=UNIT)
                rr1 = rr.registers
                rpowergraph.append(rr1)
                TextFile.write("      End of loop reflected power measurement: " + str(rr1[0]) + " Watts\n")

                try:
                    rr3 = client.read_holding_registers(102, 1, unit=UNIT)
                    rr2 = rr3.registers
                    powergraph.append(rr2)
                    TextFile.write("      End of loop transmitted power measurement: " + str(rr2[0]) + " Watts\n")

                except:
                    # sg.popup("At the end of the function, the generators power value could not be measured.\nPress Okay to continue. The code will continue without error. 0 will be appended to power list")
                    powergraph.append([0])

                # Attendre pour le temps restant de l'iteration.
                temps_restant = ONdelay - TempsTot - delaimesure

                if temps_restant <= 0:
                    pass

                else:
                    time.sleep(temps_restant)

        # ------------------------- END OF TEST, DEFAULT END STATES--------------------------------
        if i == 6:

            TextFile = open("D:\Ablation_Automatisation\Programmation\Automatisation_Andre\Logs/" + str(LogFileName) + ".txt", "a")

            # Trigger final S11 value at the end of the code
            # Ajoutée valeurs de temps en temps réel
            TempsMtn1 = str(datetime.datetime.now(pytz.timezone('America/Moncton')))
            timenow1 = TempsMtn1[11:22]
            TextFile.write("\nEnd of code time:\n" + str(timenow1) + "\n\n")
            DonneesTemps.append(timenow1)

            # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .s1p (REAL IMAGINARY DATA FORMAT)
            analyzer.write("SENS1:SWE:MODE SINGLE")
            analyzer.write("TRIGger:SCOPe CURRent")
            analyzer.write("INITiate1:IMMediate")

            # DELAY OF INPUT GUI VALUE SECONDS
            time.sleep(delaimesure)

            # TAKE MEASUREMENT AND SAVE FILE OF PORT 1 AS A .S1P (REAL IMAGINARY DATA FORMAT)
            analyzer.write("CALC:MEAS:DATA:SNP:PORTs:Save '1,,', '" + filename + "/Iteration_" + str(len(DonneesTemps)) + ".s1p'")
            analyzer.write("*OPC?")
            print("     End of code measure triggered and saved\n")
            TextFile.write("\n      End of code measure triggered and saved\n")

            rr = client.read_holding_registers(103, 1, unit=UNIT)
            rr1 = rr.registers
            rpowergraph.append(rr1)
            TextFile.write("      End of code reflected power measurement: " + str(rr1[0]) + " Watts\n")

            rr = client.read_holding_registers(102, 1, unit=UNIT)
            rr2 = rr.registers
            powergraph.append(rr2)
            TextFile.write("      End of code transmitted power measurement: " + str(rr2[0]) + " Watts\n")

            # Update next step text color on GUI window
            window.find_element('_FIVEIT5_').update(text_color='black')
            window.find_element('_FIVEIT6_').update(text_color='black')
            window.find_element('_FIVEIT7_').update(text_color='red')
            window.find_element('_FIVEIT8_').update(text_color='black')
            window.find_element('_FIVEIT9_').update(text_color='black')
            sg.OneLineProgressMeter('Test progress...', num_its + 3, num_its + 4, key='METER1',
                                    grab_anywhere=True)

            # MICROWAVES OFF
            client.write_register(2, 0x00, unit=UNIT)

            # Place switch in position I (AB)
            # rb.switchon(switch1)
            # rb.switchoff(switch2)

            # TURN PERISTALTIC PUMP OFF FOR ITS DEFAULT END STATE
            if Peris_ON == 1:

                voltageOutput0.setDeviceSerialNumber(589734)
                voltageOutput0.setChannel(0)
                voltageOutput1.setDeviceSerialNumber(589734)
                voltageOutput1.setChannel(1)

                voltageOutput0.openWaitForAttachment(1000)
                voltageOutput1.openWaitForAttachment(1000)

            
                # SET RPM SPEED AND TURN ON PUMP
                # CH1 = SPEED CONTROL CONFIG
                # CH0 = TURN PERISTALTIC PUMP OFF
                voltageOutput0.setVoltage(5)
                voltageOutput1.setVoltage(0)
                TextFile.write("\nPeristaltic pump turned OFF\n")
                print("Peristaltic pump turned OFF\n")

            # PUT ISOCRATIC PUMP IN STANDBY STATE FOR ITS DEFAULT END STATE
            if Iso_ON == 1:
                os.system("C:/Windows\Licop\LicopDemo\PumpSTANDBY1.exe")

            # Update next step text color on GUI window
            window.find_element('_FIVEIT7_').update(text_color='black')
            sg.OneLineProgressMeter('Test progress...', num_its + 4, num_its + 4, key='METER1',
                                    grab_anywhere=True)


            if Timer_ON == 1:

                # Commencer chronomètre de fin de test
                print("\nStart timer\n\n")

                EndCount = 45

                while EndCount > 0:
                    EndCount -= 1
                    print('     Time remaining on timer (s): ' + str(EndCount), end='')
                    # TextFile.write('    Time remaining on timer (s): ' + str(EndCount))
                    time.sleep(1)
                    print('\r', end='')

            # Print end of test message
            print( "\n\n------- The test is complete! Select another test or press QUIT from the drop down menu to exit. ------- \n\n")

            client.close()
            TextFile.close()



"""
S1P AVERAGE FUNCTION
"""
def S_Averages(directory, datapoints, DonneesTemps, ecart_type_reel, ecart_type_im):

    for j in range(len(DonneesTemps)):

        reel = [0] * int(datapoints)
        im = [0] * int(datapoints)
        f = [0] * int(datapoints)

        # Ouvrir et extraire donnees de parametres S11 du fichier S1P
        S1PFile = open("D:/" + directory + "/Iteration_" + str(j + 1) + ".s1p", "r")

        Header1 = S1PFile.readline()
        Header2 = S1PFile.readline()
        Header3 = S1PFile.readline()
        Header4 = S1PFile.readline()
        Header5 = S1PFile.readline()

        for __ in range(len(reel)):

            Data = S1PFile.readline().split(' ')
            f[__] = float(Data[0])
            reel[__] = float(Data[1])
            im[__] = float(Data[2])
            # S1PFile.close()

        count1 = 0
        count2 = 0
        count3 = 0

        # Calcul de la valeur moyenne de chacune des variables
        for k in range(len(f)):
            count1 += f[k]
            count2 += reel[k]
            count3 += im[k]


        av1 = count1 / len(f)
        av2 = count2 / len(reel)
        av3 = count3 / len(im)


        # Calculer le rapport d'écart-type des valeurs réels et imaginaires
        et_reel = statistics.stdev(reel) / av2
        et_im = statistics.stdev(im) / av3

        # Ajouter la valeur moyenne calculé à sa liste respective
        ecart_type_reel.append(abs(et_reel)*100)
        ecart_type_im.append(abs(et_im)*100)

        av1_corr = round(av1, 0)
        av2_corr = round(av2, 8)
        av3_corr = round(av3, 8)


        if j != 0:

            AppendFile = open("D:\Ablation_Automatisation\Programmation\Automatisation_Andre\DonneesMoyennes/" + directory + ".s1p", "a")
            AppendFile.write(str(av1_corr) + " " + str(av2_corr) + " " + str(av3_corr) + " " + "\n")

        else:
            NewFile = open("D:\Ablation_Automatisation\Programmation\Automatisation_Andre\DonneesMoyennes/" + directory + ".s1p", "w")
            NewFile.write(str(Header1))
            NewFile.write(str(Header2))
            NewFile.write(str(Header3))
            NewFile.write(str(Header4))
            NewFile.write(str(Header5))
            NewFile.write(str(av1_corr) + " " + str(av2_corr) + " " + str(av3_corr) + " " + "\n")
            NewFile.close()



"""
EXCEL DATA FORMAT FUNCTION
"""
def Excel_Data_Format(DonneesTemps, rpowergraph, powergraph, ecart_type_reel, ecart_type_im, directory):

    # Déterminer les valeurs de temps à insérer dans le fichier Excel
    rpowergraphexcel = []
    powergraphexcel = []
    TempsExcel = []
    reel = []
    im = []

    # Créer liste de données de temps
    for i in range(len(DonneesTemps)):

        if i == 0:
            TempsExcel.append(0)

        else:
            if float(DonneesTemps[i][6:11]) - float(DonneesTemps[i - 1][6:11]) < 0:
                temps_iter = float(DonneesTemps[i][6:11]) + (60 - float(DonneesTemps[i - 1][6:11]))
                TempsExcel.append(float(TempsExcel[i - 1]) + temps_iter)

            else:
                temps_iter = float(DonneesTemps[i][6:11]) - float(DonneesTemps[i - 1][6:11])
                TempsExcel.append(float(TempsExcel[i - 1]) + temps_iter)

    for i in range(len(powergraph)):
        powergraphexcel.append(powergraph[i][0])
        rpowergraphexcel.append(rpowergraph[i][0])

    """
    num_iter1 = ONDelay[0] / (delaimesure + delaimicro)
    if Flow[0] != -1:
        for j in range(int(num_iter1)):
            powergraphexcel.append(powergraph[0])
            rpowergraphexcel.append(rpowergraph[0][0])

    num_iter2 = ONDelay[1] / (delaimesure + delaimicro)
    if Flow[1] != -1:
        for j in range(int(num_iter2)):
            powergraphexcel.append(powergraph[1])
            rpowergraphexcel.append(rpowergraph[1][0])

    num_iter3 = ONDelay[2] / (delaimesure + delaimicro)
    if Flow[2] != -1:
        for j in range(int(num_iter3)):
            powergraphexcel.append(powergraph[2])
            rpowergraphexcel.append(rpowergraph[2][0])

    num_iter4 = ONDelay[3] / (delaimesure + delaimicro)
    if Flow[3] != -1:
        for j in range(int(num_iter4)):
            powergraphexcel.append(powergraph[3])
            rpowergraphexcel.append(rpowergraph[3][0])

    num_iter5 = ONDelay[4] / (delaimesure + delaimicro)
    if Flow[4] != -1:
        for j in range(int(num_iter5)):
            powergraphexcel.append(powergraph[4])
            rpowergraphexcel.append(rpowergraph[4][0])

    print(powergraphexcel)
    print(rpowergraphexcel)
    """


    # Ouvrir fichier excel et exporter valeurs
    wb = openpyxl.load_workbook(filename='D:\Ablation_Automatisation\Programmation\Automatisation_Andre\DonneesMoyennes/' + directory + '.xlsx')
    sheet_ranges = wb[directory]

    for j in range(len(DonneesTemps)):
        reel.append(sheet_ranges['B' + str(14 + j)].value)
        im.append(sheet_ranges['C' + str(14 + j)].value)

    # Placer valeurs calculés sur le fichier excel
    sheet_ranges.cell(row=13, column=2).value = 'Temps (s)'
    sheet_ranges.cell(row=13, column=3).value = 'Permitivite dielectrique reel'
    sheet_ranges.cell(row=13, column=4).value = 'Permitivite dielectrique im'
    sheet_ranges.cell(row=13, column=5).value = 'Puissance transmise (W)'
    sheet_ranges.cell(row=13, column=6).value = 'Puissance réfléchie (W)'
    sheet_ranges.cell(row=13, column=7).value = 'Écart-Type Réel (%)'
    sheet_ranges.cell(row=13, column=8).value = 'Écart-Type Im (%)'
    sheet_ranges.cell(row=13, column=9).value = 'Écart-Type Réel (eps prime)'
    sheet_ranges.cell(row=13, column=10).value = 'Écart-Type Im (eps prime prime)'


    for k in range(len(DonneesTemps)):
        sheet_ranges.cell(row=14+k, column=2).value = float(TempsExcel[k])
        sheet_ranges.cell(row=14+k, column=3).value = float(reel[k])
        sheet_ranges.cell(row=14+k, column=4).value = float(im[k])
        sheet_ranges.cell(row=14+k, column=5).value = int(powergraphexcel[k])
        sheet_ranges.cell(row=14+k, column=6).value = int(rpowergraphexcel[k])
        sheet_ranges.cell(row=14+k, column=7).value = float(ecart_type_reel[k])
        sheet_ranges.cell(row=14+k, column=8).value = float(ecart_type_im[k])
        sheet_ranges.cell(row=14+k, column=9).value = float(ecart_type_reel[k] * reel[k] / 100)
        sheet_ranges.cell(row=14+k, column=10).value = float(ecart_type_im[k] * im[k] / 100)

    wb.save(filename='D:\Ablation_Automatisation\Programmation\Automatisation_Andre\DonneesMoyennes/' + directory + '.xlsx')


"""
DIELECTRIC DATA VERIFICATION FUNCTION
"""
# A CHANGER, VALEURS DE PARAMETRES S11 ASSOCIER AVEC PARAMETRES DIELECTRIQUES
def Dielec_Data_Verif(dielec_value, datapoints, directory):

    reel = [0] * int(datapoints)
    count = 0

    # Ouvrir fichier .s1p generer recemment et extraire valeurs dielectriques reels
    # Open recent .s1p file and extract real dielectric values
    S1PFile = open("D:/" + directory + "/Iteration_" + str(i) + ".s1p", "r")

    
    for __ in range(len(reel)):

        if __ == 0:
            for i in range(5):
                DumpMeasurement = S1PFile.readline().split(' ')

        Data = S1PFile.readline().split(' ')
        reel[__] = float(Data[1])

        if dielec_value == 20 and __ == 0:
            #Valeur a changer
            s11_value = 0.75

        if dielec_value == 25 and __ == 0:
            #Valeur a changer
            s11_value = 0.7

        if dielec_value == 30 and __ == 0:
            #Valeur a changer
            s11_value = 0.65

        if dielec_value == 35 and __ == 0:
            #Valeur a changer
            s11_value = 0.6
        
        if reel[__] > s11_value:
            count += 1
    
    count_percentage = count / datapoints

    # Si au moins 25% des valeurs dielectriques sont en-dessous de la valeur desire, arrete code
    if count_percentage >= 0.25:
        break_code = 1
        sg.popup("Minimum dielectric value achieved.\nThe code will stop by itself and the GUI will keep running.")
        return break_code

    else:
        break_code = 0
        return break_code
    

"""
MANUAL S1P AVERAGE FUNCTION
"""
def S_Averages_Manu(directory, datapoints, NombreFichiers, ecart_type_reel, ecart_type_im):

        for j in range(NombreFichiers):

            reel = [0] * int(datapoints)
            im = [0] * int(datapoints)
            f = [0] * int(datapoints)

            # Ouvrir et extraire donnees de parametres S11 du fichier S1P
            S1PFile = open(directory + "/Iteration_" + str(j + 1) + ".s1p", "r")

            Header1 = S1PFile.readline()
            Header2 = S1PFile.readline()
            Header3 = S1PFile.readline()
            Header4 = S1PFile.readline()
            Header5 = S1PFile.readline()

            for __ in range(len(reel)):

                Data = S1PFile.readline().split(' ')
                f[__] = float(Data[0])
                reel[__] = float(Data[1])
                im[__] = float(Data[2])
                # S1PFile.close()

            count1 = 0
            count2 = 0
            count3 = 0

            # Calcul de la valeur moyenne de chacune des variables
            for k in range(len(f)):
                count1 += f[k]
                count2 += reel[k]
                count3 += im[k]


            av1 = count1 / len(f)
            av2 = count2 / len(reel)
            av3 = count3 / len(im)


            # Calculer le rapport d'écart-type des valeurs réels et imaginaires
            et_reel = statistics.stdev(reel) / av2
            et_im = statistics.stdev(im) / av3

            # Ajouter la valeur moyenne calculé à sa liste respective
            ecart_type_reel.append(abs(et_reel)*100)
            ecart_type_im.append(abs(et_im)*100)

            av1_corr = round(av1, 0)
            av2_corr = round(av2, 8)
            av3_corr = round(av3, 8)


            if j != 0:

                AppendFile = open(directory + "/DonneesMoyennes.s1p", "a")
                AppendFile.write(str(av1_corr) + " " + str(av2_corr) + " " + str(av3_corr) + " " + "\n")

                StdDevFile = open(directory + "/StdDev_Values.txt", "a")
                StdDevFile.write(str(ecart_type_reel[j]) + " " + str(ecart_type_im[j]) + "\n")

            else:
                NewFile = open(directory + "/DonneesMoyennes.s1p", "w")
                StdDevFile = open(directory + "/StdDev_Values.txt", "w")
                StdDevFile.write(str(ecart_type_reel[j]) + " " + str(ecart_type_im[j]) + "\n")
                NewFile.write(str(Header1))
                NewFile.write(str(Header2))
                NewFile.write(str(Header3))
                NewFile.write(str(Header4))
                NewFile.write(str(Header5))
                NewFile.write(str(av1_corr) + " " + str(av2_corr) + " " + str(av3_corr) + " " + "\n")
                NewFile.close()
                StdDevFile.close()



"""
**********************************************
**************** UI INTERFACE ****************
**********************************************
"""

sg.ChangeLookAndFeel('BlueMono')

# MENU DEFINITION
menu_def = [['OPTIONS', ['CREATE', 'QUIT']]]

# INITIAL WINDOWS LAYOUTS
Init_layout = [
    [sg.Text('Welcome to the automation circuit UI!', font=("Helvetica", 20), pad=(270, 5))],
    [sg.Text('Please select one of the buttons to get started', font=("Helvetica", 20), pad=(230, 5))],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\IARC_LOGO.png',
        pad=(100, 0)),
     sg.Image(
         'D:\Ablation_Automatisation\Programmation\Images\LOGO_UdeM.png',
         pad=(100, 0))],
    [sg.Text('Parameter configuration', font=("Helvetica", 18), pad=(370, 5))],
    [sg.Submit(key='_ALLPARAMETERS_OK_', button_text='Select', font=("Helvetica", 18), pad=(440, 20))],
    [sg.Text('Tests and manual component control', font=("Helvetica", 18), pad=(300, 5)), ],
    [sg.Submit(key='_TEST_OK_', font=("Helvetica", 18), button_text='Select', pad=(440, 20),
               auto_size_button=True)]]

Parameter_layout = [
    [sg.Text('Network Analyzer', font=("Helvetica", 12), pad=(30, 5)), 
     sg.Text('Generator', font=("Helvetica", 12), pad=(170, 0)),
     sg.Text('Pumps', font=("Helvetica", 12), pad=(170, 0))],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\ENA.png',
        pad=(50, 5)), 
     sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\Generatrice.png',
        pad=(50, 5)),
     sg.Text('', font=("Helvetica", 12), pad=(75, 5)),
     sg.Image(
         'D:\Ablation_Automatisation\Programmation\Images\IsoPump.png',
         pad=(0, 5)),
     sg.Image(
         'D:\Ablation_Automatisation\Programmation\Images\PerisPump.png',
         pad=(0, 5))],
    [sg.Submit(key='_ANA_OK_', button_text='Select', font=("Helvetica", 12), pad=(100, 0)),
     sg.Submit(key='_GEN_OK_', font=("Helvetica", 12), button_text='Select', pad=(100, 0), auto_size_button=True),
     sg.Submit(key='_PERIS_OK_', button_text='Select', font=("Helvetica", 12), pad=(280, 0))],
    [sg.Cancel(key='_ALLPARAMETERS_CANCEL_', font=("Helvetica", 12), button_text='Return', pad=(0, 0),
               auto_size_button=True)]]

Test_layout = [
    [sg.Text('TESTS', font=("Helvetica", 16), pad=(0, 5))],
    [sg.Text('', font=("Helvetica", 5), pad=(0, 0))],
    [sg.Text('                                                 Normal test', font=("Helvetica", 14),
             pad=(20, 5)),
     sg.Text('Five iteration test', font=("Helvetica", 14), pad=(30, 5)),
     sg.Text('Sweep test', font=("Helvetica", 14), pad=(0, 5))],
    [sg.Text('', font=("Helvetica", 14), pad=(150, 0)),
     sg.Submit(key='_LAUNCH_OK_', button_text='Launch', font=("Helvetica", 14), pad=(40, 5)),
     sg.Submit(key='_TEST_OK3_', button_text='Select', font=("Helvetica", 14), pad=(40, 5)),
     sg.Submit(key='_TEST_OK1_', button_text='Select', font=("Helvetica", 14), pad=(40, 5))],
    [sg.Text('MANUAL COMPONENT CONTROL', font=("Helvetica", 16), pad=(0, 20))],
    [sg.Text('                                                  Generator', font=("Helvetica", 14),
             pad=(20, 5)),
     sg.Text('Network Analyzer', font=("Helvetica", 14), pad=(30, 5)),
     sg.Text('Pumps', font=("Helvetica", 14), pad=(30, 5)),
     sg.Text('S1P Average', font=("Helvetica", 14), pad=(0, 5))],
    [sg.Text('', font=("Helvetica", 14), pad=(150, 0)),
     sg.Submit(key='_TEST_OK2_', button_text='Select', font=("Helvetica", 14), pad=(40, 5)),
     sg.Submit(key='_TEST_OK4_', button_text='Select', font=("Helvetica", 14), pad=(40, 5)),
     sg.Submit(key='_TEST_OK5_', button_text='Select', font=("Helvetica", 14), pad=(40, 5)),
     sg.Submit(key='_TEST_OK6_', button_text='Select', font=("Helvetica", 14), pad=(40, 5))],
    [sg.Cancel(key='_TEST_CANCEL_', font=("Helvetica", 14), button_text='Return', pad=(0, 10), auto_size_button=True)]]

# GENERATOR SETTINGS LAYOUT
gen_layout = [
    [sg.Text('Microwave power (W, insert 0 if not configuring)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_POWER1_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('ON delay duration (s, insert 0 if not configuring)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ONDELAY1_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('OFF delay duration (s, insert 0 if not configuring)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_OFFDELAY1_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Frequency (MHz)', font=("Helvetica", 18)),
     sg.Input('2450', do_not_clear=True, key='_FREQ1_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Reflected power (W, set to 0 for automatic configuration (50%))', font=("Helvetica", 18), text_color='red',
             background_color='yellow'),
     sg.Input('0', do_not_clear=True, key='_REFPOWER_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Submit(key='_SET_GEN_', font=("Helvetica", 18), button_text='Set other values', pad=(170, 10),
               auto_size_button=True),
     sg.Submit(key='_SET_RPOWER_', font=("Helvetica", 18), button_text='Set reflected power', pad=(170, 10),
               auto_size_button=True)],
    [sg.Cancel(key='_GEN_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0), auto_size_button=True)]]

# NETWORK ANALYZER SETTINGS LAYOUT
Analyzer_layout = [
    [sg.Text('Frequencies, datapoints & sweep mode', font=("Helvetica", 18), pad=(0, 5)), ],
    [sg.Submit(key='_FREQ_OK_', button_text='Select', font=("Helvetica", 18), pad=(300, 20))],
    [sg.Text('# of iterations & filename', font=("Helvetica", 18), pad=(0, 5)), ],
    [sg.Submit(key='_PARAMS1_OK_', button_text='Select', font=("Helvetica", 18), pad=(300, 20))],
    [sg.Cancel(key='_ANA_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0), auto_size_button=True)]]

Freq_layout = [
    [sg.Text('Start frequency (MHz)', font=("Helvetica", 18)),
     sg.Input('2450', do_not_clear=True, key='_STARTFREQ_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Stop frequency (MHz)', font=("Helvetica", 18)),
     sg.Input('2450', do_not_clear=True, key='_STOPFREQ_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Number data points ', font=("Helvetica", 18)),
     sg.Input('101', do_not_clear=True, key='_NUMDATA_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Bandwidth (Hz)', font=("Helvetica", 18)),
     sg.Input('300', do_not_clear=True, key='_BW_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Radio('LOG SWEEP', "RADIO5", default=False, change_submits=True, key="_LOG_ON_", font=("Helvetica", 18),
              size=(15, 1), pad=(75, 10)),
     sg.Radio('LINEAR SWEEP', "RADIO5", default=True, change_submits=True, key="_LOG_OFF_", font=("Helvetica", 18),
              size=(15, 1), pad=(75, 10))],
    [sg.Submit(key='_CONFIG2_OK_', font=("Helvetica", 18), button_text='Set values', pad=(300, 10),
               auto_size_button=True)],
    [sg.Cancel(key='_FREQ_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0), auto_size_button=True)]]

Params1_layout = [
    [sg.Text('Number of test iterations (insert 0 if not configuring)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_NUMITS_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Submit(key='_CONFIG1_OK_', font=("Helvetica", 18), button_text='Set values', pad=(300, 10),
               auto_size_button=True)],
    [sg.Cancel(key='_PARAMS1_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0),
               auto_size_button=True)]]

# PERISTALTIC PUMP LAYOUT
Peristaltic_layout = [
    [sg.Text('Cooling flow speed (RPM - set to 0 if not using)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_RPM_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Radio('COOLING ON', "RADIO2", default=False, change_submits=True, key="_PERIS_ON_", font=("Helvetica", 18),
              size=(15, 1), pad=(50, 10)),
     sg.Radio('COOLING OFF', "RADIO2", default=False, change_submits=True, key="_PERIS_OFF_",
              font=("Helvetica", 18), size=(15, 1), pad=(50, 10))],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\PerisPump.png',
        pad=(50, 10))],
    [sg.Radio('ALCOHOL ON', "RADIO3", default=False, change_submits=True, key="_ISO_ON_", font=("Helvetica", 18),
              size=(15, 1), pad=(50, 0)),
     sg.Radio('ALCOHOL OFF', "RADIO3", default=False, change_submits=True, key="_ISO_OFF_",
              font=("Helvetica", 18), size=(15, 1), pad=(50, 0))],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\IsoPump.png',
        pad=(50, 10))],
    [sg.Submit(key='_CONFIG7_OK_', font=("Helvetica", 18), button_text='Set values', pad=(300, 10),
               auto_size_button=True)],
    [sg.Cancel(key='_PERIS_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0), auto_size_button=True)]]

# TEST - Frequency sweep scan selection layout
sweep_layout = [
    [sg.Text("Description: For this test, the ENA and the generators' frequencies are adjusted at each iteration",
             font=("Helvetica", 18))],
    [sg.Text("", font=("Helvetica", 5))],
    [sg.Text('Step frequency (MHz , min. 100 KHz, max. 10 MHz): ', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_STEPFREQ_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Start frequency (MHz)', font=("Helvetica", 18)),
     sg.Input('2400', do_not_clear=True, key='_STARTFREQ1_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Stop frequency (MHz)', font=("Helvetica", 18)),
     sg.Input('2500', do_not_clear=True, key='_STOPFREQ1_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Radio('Manual mode', "RADIO1", default=False, change_submits=True, key="_MANU_MODE_", font=("Helvetica", 18),
              size=(10, 1), pad=(100, 10)),
     sg.Radio('Auto mode', "RADIO1", default=False, change_submits=True, key="_AUTO_MODE_", font=("Helvetica", 18),
              size=(10, 1), pad=(100, 10))],
    [sg.Submit(key='_CONFIG4_OK_', font=("Helvetica", 18), button_text='Configure', pad=(100, 10),
               auto_size_button=True),
     sg.Submit(key='_TEST2_OK_', font=("Helvetica", 18), button_text='Launch test', pad=(100, 10),
               auto_size_button=True)],
    [sg.Cancel(key='_SWEEP_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0), auto_size_button=True)],
    [sg.Text("", font=("Helvetica", 5))],
    [sg.Text('TEST EXECUTIONS', font=("Helvetica", 18), text_color='green', key='_SWEEP1_', pad=(10, 10))],
    [sg.Text('Initialisations', font=("Helvetica", 18), key='_SWEEP2_', pad=(10, 0))],
    [sg.Text('Dump first measurement', font=("Helvetica", 18), key='_SWEEP3_', pad=(10, 0))],
    [sg.Text('FOR LOOP', font=("Helvetica", 18), key='_SWEEP4_', pad=(10, 0))],
    [sg.Text('   Microwaves ON', font=("Helvetica", 18), key='_SWEEP5_', pad=(10, 0))],
    [sg.Text('   Microwaves OFF / Dielectric measurement', font=("Helvetica", 18), key='_SWEEP6_', pad=(10, 0))],
    [sg.Text('Test complete', font=("Helvetica", 18), key='_SWEEP7_', pad=(10, 0))],
    [sg.Text('', font=("Helvetica", 14), pad=(0, 0))],
    [sg.Text('PUMPS', font=("Helvetica", 14), text_color='green', pad=(0, 0))],
    [sg.Text('Alcohol/Isocratic pump ON', font=("Helvetica", 14), key='_SWEEP8_', pad=(0, 0))],
    [sg.Text('Cooling/Peristaltic pump ON', font=("Helvetica", 14), key='_SWEEP9_', pad=(0, 0))]]

# TEST - Normal text layout
normal_text_layout = [
    [sg.Text('NORMAL TEST EXECUTIONS', font=("Helvetica", 14), text_color='green', key='_NORMAL1_', pad=(0, 10))],
    [sg.Text('Initialisations', font=("Helvetica", 14), key='_NORMAL2_', pad=(0, 0))],
    [sg.Text('Dump first measurement', font=("Helvetica", 14), key='_NORMAL3_', pad=(0, 0))],
    [sg.Text('FOR LOOP', font=("Helvetica", 14), key='_NORMAL4_', pad=(0, 0))],
    [sg.Text('   Microwaves OFF / Dielectric measurement', font=("Helvetica", 14), key='_NORMAL5_', pad=(0, 0))],
    [sg.Text('   Microwaves ON', font=("Helvetica", 14), key='_NORMAL6_', pad=(0, 0))],
    [sg.Text('Test complete', font=("Helvetica", 14), key='_NORMAL7_', pad=(0, 0))],
    [sg.Text('', font=("Helvetica", 14), pad=(0, 0))],
    [sg.Text('PUMPS', font=("Helvetica", 14), text_color='green', pad=(0, 0))],
    [sg.Text('Alcohol/Isocratic pump ON', font=("Helvetica", 14), key='_NORMAL8_', pad=(0, 0))],
    [sg.Text('Cooling/Peristaltic pump ON', font=("Helvetica", 14), key='_NORMAL9_', pad=(0, 0))]]

# TEST - Manual generator layout
manu_gen_layout = [
    [sg.Text('Turn ON generator (freq. = 2.45GHz)', font=("Helvetica", 18)),
     sg.Submit(key='_ONSET_OK_', font=("Helvetica", 18), button_text='GEN ON', pad=(210, 0), auto_size_button=True)],
    [sg.Text('Power (W)', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_POWER3_', size=(10, 10), font=("Helvetica", 18)),
     sg.Submit(key='_OFFSET_OK_', font=("Helvetica", 18), button_text='GEN OFF', pad=(350, 0), auto_size_button=True)],
    [sg.Text('Reflected power (W, set to 0 for automatic configuration)', font=("Helvetica", 18), text_color='red',
             background_color='yellow'),
     sg.Input('0', do_not_clear=True, key='_RPOWER_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('', font=("Helvetica", 18))],
    [sg.Text('STATUS', text_color='green', font=("Helvetica", 18))],
    [sg.Text('Generator OFF', font=("Helvetica", 18), text_color='red', key='OFFSETTEXT')],
    [sg.Text('Generator ON', font=("Helvetica", 18), text_color='black', key='ONSETTEXT')],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\Generatrice.png',
        pad=(10, 0))],
    [sg.Cancel(key='_MANUGEN_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0),
               auto_size_button=True)]]

# TEST - Manual ENA layout
manu_ENA_layout = [
    [sg.Text('Start Frequency (MHz)', font=("Helvetica", 18)),
     sg.Input('2440', do_not_clear=True, key='_STARTFREQMANU_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Stop Frequency (GHz)', font=("Helvetica", 18)),
     sg.Input('2460', do_not_clear=True, key='_STOPFREQMANU_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('# of datapoints', font=("Helvetica", 18)),
     sg.Input('101', do_not_clear=True, key='_DATAPOINTSMANU_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Bandwidth (Hz)', font=("Helvetica", 18)),
     sg.Input('300', do_not_clear=True, key='_BWMANU_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('File directory', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_DIRECTORYMANU_', size=(30, 10), font=("Helvetica", 18))],
    [sg.Text('Measurement delay', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_DELAYENAMANU_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Radio('LOG SWEEP', "RADIO6", default=False, change_submits=True, key="_LOG_ON1_", font=("Helvetica", 18),
              size=(15, 1), pad=(75, 10)),
     sg.Radio('LINEAR SWEEP', "RADIO6", default=True, change_submits=True, key="_LOG_OFF1_", font=("Helvetica", 18),
              size=(15, 1), pad=(75, 10))],
    [sg.Submit(key='_CONFIG6_OK_', font=("Helvetica", 18), button_text='Set values', pad=(100, 10),
               auto_size_button=True),
     sg.Submit(key='_TEST3_OK_', font=("Helvetica", 18), button_text='Trigger ENA', pad=(100, 10),
               auto_size_button=True)],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\ENA.png',
        pad=(250, 0))],
    [sg.Cancel(key='_MANUENA_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0),
               auto_size_button=True)]]

# TEST - Manual Pumps layout
manu_pumps_layout = [
    [sg.Text('Cooling flow speed (RPM, must not leave blank)', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_RPM1_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Radio('COOLING ON', "RADIO4", default=False, change_submits=True, key="_MANU_PERIS_ON_", font=("Helvetica", 18),
              size=(15, 1), pad=(50, 10)),
     sg.Radio('COOLING OFF', "RADIO4", default=False, change_submits=True, key="_MANU_PERIS_OFF_",
              font=("Helvetica", 18), size=(15, 1), pad=(50, 10))],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\PerisPump.png',
        pad=(50, 0))],
    [sg.Radio('ALCOHOL ON', "RADIO4", default=False, change_submits=True, key="_MANU_ISO_ON_", font=("Helvetica", 18),
              size=(15, 1), pad=(50, 10)),
     sg.Radio('ALCOHOL OFF', "RADIO4", default=False, change_submits=True, key="_MANU_ISO_OFF_",
              font=("Helvetica", 18), size=(15, 1), pad=(50, 10))],
    [sg.Image(
        'D:\Ablation_Automatisation\Programmation\Images\IsoPump.png',
        pad=(50, 0))],
    [sg.Text('', font=("Helvetica", 10))],
    [sg.Text('STATUS', text_color='green', font=("Helvetica", 18))],
    [sg.Text('Cooling OFF', font=("Helvetica", 18), text_color='red', key='_COOLINGOFF_')],
    [sg.Text('Cooling ON', font=("Helvetica", 18), text_color='black', key='_COOLINGON_'),
     sg.Submit(key='_TEST4_OK_', font=("Helvetica", 18), button_text='Launch', pad=(300, 0), auto_size_button=True)],
    [sg.Text('Alcohol OFF', font=("Helvetica", 18), text_color='red', key='_ALCOHOLOFF_')],
    [sg.Text('Alcohol ON', font=("Helvetica", 18), text_color='black', key='_ALCOHOLON_')],
    [sg.Text('', font=("Helvetica", 10))],
    [sg.Cancel(key='_MANUPUMPS_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0),
               auto_size_button=True)]]

# TEST - Five iteration test layout
five_it_layout = [
    [sg.Text('Description: For this test, the power and delay durations can be changed at each iteration. An iteration can be skipped by inserting -1 in the drop down menu',
             font=("Helvetica", 12))],
    [sg.Text("", font=("Helvetica", 10))],
    [sg.Text('Power iteration 1 (W)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ITPOWER1_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Duration (s): ', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ONDELAY31_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Pump flow rate (ml/min)', font=("Helvetica", 18)),
     sg.Combo(values=['-1', '0', '1', '2', '2.5', '3', '4', '5', '10'], key='_LISTBOX1_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Power iteration 2 (W)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ITPOWER2_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Duration (s): ', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ONDELAY32_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Pump flow rate (ml/min)', font=("Helvetica", 18)),
     sg.Combo(values=['-1', '0', '1', '2', '2.5', '3', '4', '5', '10'], key='_LISTBOX2_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Power iteration 3 (W)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ITPOWER3_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Duration (s): ', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ONDELAY33_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Pump flow rate (ml/min)', font=("Helvetica", 18)),
     sg.Combo(values=['-1', '0', '1', '2', '2.5', '3', '4', '5', '10'], key='_LISTBOX3_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Power iteration 4 (W)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ITPOWER4_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Duration (s): ', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ONDELAY34_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Pump flow rate (ml/min)', font=("Helvetica", 18)),
     sg.Combo(values=['-1', '0', '1', '2', '2.5', '3', '4', '5', '10'], key='_LISTBOX4_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Power iteration 5 (W)', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ITPOWER5_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Duration (s): ', font=("Helvetica", 18)),
     sg.Input('0', do_not_clear=True, key='_ONDELAY35_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Pump flow rate (ml/min)', font=("Helvetica", 18)),
     sg.Combo(values=['-1', '0', '1', '2', '2.5', '3', '4', '5', '10'], key='_LISTBOX5_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Délai premier état (Micro Onde ON)', font=("Helvetica", 18)),
     sg.Input('5', do_not_clear=True, key='_DELAIMICRO_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Délai deuxième état (Mesure diélectrique)', font=("Helvetica", 18)),
     sg.Input('1', do_not_clear=True, key='_DELAIMESURE_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Desired saved filename', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_DIRECTORY_', size=(20, 30), font=("Helvetica", 18), pad=(20, 0)),
     sg.Text('Dielectric verification?', font=("Helvetica", 18)),
     sg.Radio('Yes', "RADIO6", default=False, change_submits=True, key="_DIELEC_VERIF_", font=("Helvetica", 18),
              size=(15, 1)),
     sg.Radio('No', "RADIO6", default=False, change_submits=True, key="_DIELEC_VERIF1_", font=("Helvetica", 18),
              size=(15, 1))],
    [sg.Text('Minimum dielectric value', font=("Helvetica", 18)),
     sg.Combo(values=['20', '25', '30', '35'], key='_LISTBOX6_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('                Timer ON?', font=("Helvetica", 18)),
     sg.Radio('Yes', "RADIO7", default=False, change_submits=True, key="_TIMER_ON_", font=("Helvetica", 18),
              size=(15, 1)),
     sg.Radio('No', "RADIO7", default=False, change_submits=True, key="_TIMER_OFF_", font=("Helvetica", 18),
              size=(15, 1))],    
    [sg.Cancel(key='_FIVEIT_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0), auto_size_button=True),
     sg.Submit(key='_CONFIG5_OK_', font=("Helvetica", 18), button_text='Configure', pad=(100, 0),
               auto_size_button=True),
     sg.Submit(key='_TEST5IT_OK_', font=("Helvetica", 18), button_text='Launch', pad=(70, 0), auto_size_button=True),     
     sg.Submit(key='_RESET_FAULTS_', font=("Helvetica", 18), button_text='Reset gen. faults', pad=(20, 0),
               auto_size_button=True),
     sg.Submit(key='_S1PAVERAGE_', font=("Helvetica", 13), button_text='S1P file average', pad=(20, 0),
               auto_size_button=True),
     sg.Submit(key='_EXCELDATA_', font=("Helvetica", 13), button_text='Excel Data', pad=(20, 0),
               auto_size_button=True)],
    [sg.Text("", font=("Helvetica", 10))],
    [sg.Text('TEST EXECUTIONS', font=("Helvetica", 18), text_color='green', key='_FIVEIT1_', pad=(10, 10))],
    [sg.Text('Initialisations', font=("Helvetica", 18), key='_FIVEIT2_', pad=(10, 0))],
    [sg.Text('Dump first measurement', font=("Helvetica", 18), key='_FIVEIT3_', pad=(10, 0))],
    [sg.Text('FOR LOOP', font=("Helvetica", 18), key='_FIVEIT4_', pad=(10, 0))],
    [sg.Text('   Microwaves ON', font=("Helvetica", 18), key='_FIVEIT5_', pad=(10, 0))],
    [sg.Text('   Microwaves OFF / Dielectric measurement', font=("Helvetica", 18), key='_FIVEIT6_', pad=(10, 0))],
    [sg.Text('Test complete', font=("Helvetica", 18), key='_FIVEIT7_', pad=(10, 0))],
    [sg.Text('PUMPS', font=("Helvetica", 14), text_color='green', pad=(0, 0))],
    [sg.Text('Alcohol/Isocratic pump ON', font=("Helvetica", 14), key='_FIVEIT8_', pad=(0, 0))],
    [sg.Text('Cooling/Peristaltic pump ON', font=("Helvetica", 14), key='_FIVEIT9_', pad=(0, 0))]]

# PROGRESS BAR LAYOUT
progress_bar_layout = [[sg.Text('Test progress meter')],
                       [sg.ProgressBar(1000, orientation='h', size=(20, 20), key='METER1')],
                       [sg.Cancel()]]

# Manual average s1p layout
manu_S1P_layout = [
    [sg.Text('Number of datapoints', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_DATAPOINTS_S1P_', size=(10, 10), font=("Helvetica", 18)),
     sg.Text('Number of iterations (inside respective data folder)', font=("Helvetica", 18)),
     sg.Input('', do_not_clear=True, key='_ITERATIONS_S1P_', size=(10, 10), font=("Helvetica", 18))],
    [sg.Text('Full directory (ex. D:\Ablation_Automatisation\Programmation)', font=("Helvetica", 18)),
     sg.Input('D:/', do_not_clear=True, key='_DIRECTORY_S1P_', size=(40, 10), font=("Helvetica", 18))],
    [sg.Submit(key='_CONFIG8_OK_', font=("Helvetica", 18), button_text='Set values', pad=(100, 10),
               auto_size_button=True),
     sg.Submit(key='_MANUS1P_OK_', font=("Helvetica", 18), button_text='S1P Average calculation', pad=(100, 10),
               auto_size_button=True)],
    [sg.Cancel(key='_MANUS1P_CANCEL_', font=("Helvetica", 18), button_text='Return', pad=(0, 0),
               auto_size_button=True)]]



'''
MAIN USER LAYOUT
'''
# Frame creation (For main window)
layout = [[sg.Menu(menu_def, tearoff=True)],
          [sg.Frame('', Init_layout, font='Any 13', title_color='blue', visible=False,
                    key='_INIT_FRAME_', size=(800, 400)),
           sg.Frame('Parameters options', Parameter_layout, font='Any 13', title_color='blue', visible=False,
                    key='_ALLPARAMETERS_FRAME_', size=(1200, 600)),
           sg.Frame('Generator parameters options', gen_layout, font='Any 13', title_color='blue', visible=False,
                    key='_GEN_FRAME_', size=(800, 300)),
           sg.Frame('ENA parameters', Analyzer_layout, font='Any 13', title_color='blue', visible=False,
                    key='_ANA_FRAME_', size=(800, 400)),
           sg.Frame('ENA parameters option 1', Freq_layout, font='Any 13', title_color='blue', visible=False,
                    key='_FREQ_FRAME_', size=(1000, 400)),
           sg.Frame('ENA parameters option 2', Params1_layout, font='Any 13', title_color='blue', visible=False,
                    key='_PARAMS1_FRAME_', size=(800, 400)),
           sg.Frame('Pumps parameters options', Peristaltic_layout, font='Any 13', title_color='blue', visible=False,
                    key='_PUMPS_FRAME_', size=(800, 400)),
           sg.Frame('Tests and manual component control options', Test_layout, font='Any 13', title_color='blue',
                    visible=False,
                    key='_TEST_FRAME_', size=(800, 400)),
           sg.Frame('Sweep test options (SELECT MANUAL MODE ONLY)', sweep_layout, font='Any 13', title_color='blue',
                    visible=False,
                    key='_SWEEP_FRAME_', size=(1200, 600)),
           sg.Frame('Manual generator control options', manu_gen_layout, font='Any 13', title_color='blue',
                    visible=False, key='_MANUGEN_FRAME_', size=(800, 400)),
           sg.Frame('Manual ENA control options', manu_ENA_layout, font='Any 13', title_color='blue', visible=False,
                    key='_MANUENA_FRAME_', size=(800, 400)),
           sg.Frame('Manual pumps control options', manu_pumps_layout, font='Any 13', title_color='blue', visible=False,
                    key='_MANUPUMPS_FRAME_', size=(800, 400)),
           sg.Frame('Five iteration test options', five_it_layout, font='Any 13', title_color='blue', visible=False,
                    key='_FIVEIT_FRAME_', size=(1200, 600)),
           sg.Frame('', normal_text_layout, font='Any 13', visible=False,
                    key='_NORMALTEXT_FRAME_', size=(400, 300)),
           sg.Frame('Manual s1p average window', manu_S1P_layout, font='Any 13', visible=False,
                    key='_S1PAVERAGE_FRAME_', size=(400, 300))
           ]]

# Main user interface window popup
window = sg.Window('AUTOMATION CODE - V5', layout, default_element_size=(40, 1), grab_anywhere=False,
                   size=(1600, 1000))

"""
***********************************************
**************** MAIN FUNCTION ****************
***********************************************
"""


def main():
    value_dict = {"startfreq": [], "stopfreq": [], "datapoints": [], "bw": [], "num_its": [], "ondelay1": [],
                  "offdelay1": [],
                  "directory": [], "switch1": [], "switch2": [], "power1": [], "freq1": [], "dpower2": [],
                  "startfreq2": [], "stopfreq2": [], "stepfreq": [], "power3": [],
                  "rpower": [], "rpower1": [], "itpower": [],
                  "ondelay3": [], "Listbox1": [], "startfreqmanu": [], "stopfreqmanu": [],
                  "datapointsmanu": [], "bwmanu": [],
                  "delayENAmanu": [], "directorymanu": [], "RPM": [], "RPM1": [], "inputpower": [], "delaimicro": [],
                  "delaimesure": [], "donneestemps": [], "rpowergraph": [], "powergraph": [], "stdevreel": [], "stdevim": [],
                  "Listbox2": [], "datapointsmanu1": [], "directorymanu1": [], "num_itsmanu": []}

    # Initialise bit trigger values
    IsManu = 1
    IsLog = 0
    IsLogManu = 0
    Peris_ON = 0
    Iso_ON = 0
    Manu_Pumps_State = 0
    Manu_Iso_ON = 0
    Manu_Peris_ON = 0
    Cool_Pump_ON = 0
    Dielec_Verif = 0
    Timer_ON = 0

    # Print introductory message
    print("\n\n------------------WELCOME TO THE AUTOMATION CIRCUIT V5 UI! Select OPTIONS -> CREATE to get started.------------------\n")

    while True:
        event, values = window.Read()

        # For debugging purposes
        # print(event)

        """
        INIT WINDOW
        """
        if event == 'CREATE':
            window.FindElement('_INIT_FRAME_').Update(visible=True)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_ALLPARAMETERS_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=True)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_TEST_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=True)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        """
        TEST RUN FUNCTIONS
        """
        # Turn generator on manually
        if event == '_ONSET_OK_':

            value_dict["power3"] = []
            value_dict["rpower"] = []

            try:

                value_dict["rpower"].append(int(values['_RPOWER_']))
                value_dict["power3"].append(int(values['_POWER3_']))

                # Clause de sécurité. Prévenir l'utilisateur que les valeurs de puissance configurés requis l'accompagnement de la pompe à refroidisement.
                if 20 <= value_dict["power3"][0] < 200:
                    sg.popup(
                        "For the values of transmitted power you've selected, cooling might/will be required.\nPress OK to continue.")

                # Clause de sécurité. Prévenir à ce que l'utilisateur choisisse une valeur de puissance au-delà de 200 W.
                if value_dict["power3"][0] >= 200:
                    sg.popup(
                        "The value of transmitted power selected is too high.\nPlease select another value. The value of transmitted power has been set to 0 Watts.\nPress OK to continue.")
                    value_dict["power3"][0] = 0

                if int(value_dict["rpower"][0]) != 0:
                    sg.popup(
                        "Are you sure you want to manually configure the reflected power?\nIts value should not exceed 15% of the transmitted powers' value.\nPress OK to continue.")

                window.FindElement('_MANUGEN_FRAME_').Update(visible=True)
                Manual_Gen_ON(value_dict['power3'][0], value_dict['rpower'][0])

            except:
                sg.popup("Please check the fields.")

        # Turn generator off manually
        if event == '_OFFSET_OK_':

            try:
                window.FindElement('_MANUGEN_FRAME_').Update(visible=True)
                Manual_Gen_OFF()

            except:
                sg.popup("Please check the fields!")

        # Frequency sweep test
        if event == '_TEST2_OK_':
            try:
                # sg.popup("Press OK to confirm launch. Press X to exit")
                Freq_sweep(value_dict['datapoints'][0],
                           value_dict['bw'][0], value_dict['directory'][0], value_dict['power1'][0],
                           value_dict['rpower1'][0],
                           value_dict['ondelay1'][0], value_dict['offdelay1'][0],
                           value_dict['startfreq1'][0], value_dict['stopfreq1'][0], value_dict['stepfreq'][0], IsManu,
                           Peris_ON, value_dict['RPM'][0], Iso_ON,
                           IsLog)

                window.FindElement('_SWEEP_FRAME_').Update(visible=True)
                sg.popup("Test is complete!")

            except:
                sg.popup('Please check the fields')
                window.FindElement('_SWEEP_FRAME_').Update(visible=True)

        # Manual ENA trigger
        if event == '_TEST3_OK_':
            try:
                # sg.popup("Press OK to confirm launch. Press X to exit")
                Manual_ENA_Trigger(value_dict['startfreqmanu'][0], value_dict['stopfreqmanu'][0],
                                   value_dict['datapointsmanu'][0],
                                   value_dict['bwmanu'][0], value_dict['delayENAmanu'][0],
                                   value_dict['directorymanu'][0], IsLogManu)

                window.FindElement('_MANUENA_FRAME_').Update(visible=True)
                sg.popup("Trigger complete!")

            except:
                sg.popup('Please check the fields')
                window.FindElement('_MANUENA_FRAME_').Update(visible=True)

        # Manual pumps trigger
        if event == '_TEST4_OK_':

            value_dict["RPM1"] = []

            try:
                value_dict["RPM1"].append(int(values['_RPM1_']))

                """
                COOLING PUMP INIT
                """
                if Cool_Pump_ON == 0:
                    voltageOutput0 = VoltageOutput()
                    voltageOutput1 = VoltageOutput()
                    Cool_Pump_ON = 1

                # Clause de sécurité. Prévenir à ce que l'utilisateur choisisse une valeur de vitesse de débit chez la pompe péristaltique au-delà de 300 RPM.
                if value_dict["RPM1"][0] >= 120:
                    sg.popup("The value of cooling speed selected is too high.\nPlease select another value. The value of speed has been set to 0 RPM.\nPress OK to continue.")
                    value_dict["RPM1"][0] = 0

                Manual_Pumps(Manu_Pumps_State, value_dict['RPM1'][0], Manu_Iso_ON, Manu_Peris_ON, voltageOutput0, voltageOutput1)

                window.FindElement('_MANUPUMPS_FRAME_').Update(visible=True)
                # window.reappear()

            except:
                sg.popup('Please check the fields')
                window.FindElement('_MANUPUMPS_FRAME_').Update(visible=True)

        # 5 iteration test
        if event == '_TEST5IT_OK_':

            # try:
            # window.disappear()
            # sg.popup("Press OK to confirm launch. Press X to exit")

            # Vérification: Si toutes les étapes ne requièrent pas l'utilisation de la pompe d'alcool, elle ne sera pas activée lors du test.
            if value_dict['Listbox1'][0] == 0 and value_dict['Listbox1'][1] == 0 and value_dict['Listbox1'][2] == 0 and value_dict['Listbox1'][3] == 0 and value_dict['Listbox1'][4] == 0 and Iso_ON == 1:
                Iso_ON = 0

            value_dict['donneestemps'] = []
            value_dict['rpowergraph'] = []
            value_dict['powergraph'] = []

            # Placer l'heure et la date actuelle sur le fichier de text
            current_time = str(datetime.datetime.now(pytz.timezone('America/Moncton')))
            corrected_time = current_time.replace(":", ".")

            """
            COOLING PUMP INIT
            """
            if Cool_Pump_ON == 0:
                voltageOutput0 = VoltageOutput()
                voltageOutput1 = VoltageOutput()
                Cool_Pump_ON = 1
            
            # Exit = 0

            for i in range(6):

                Exit = Five_Iteration_Test(value_dict['startfreq'][0], value_dict['stopfreq'][0],
                                    value_dict['datapoints'][0],
                                    value_dict['bw'][0], value_dict['directory'][0], value_dict['delaimicro'][0],
                                    value_dict['delaimesure'][0], voltageOutput0, voltageOutput1, value_dict['itpower'][i - 1],
                                    value_dict['ondelay3'][i - 1], value_dict['Listbox1'][i - 1],
                                    value_dict['freq1'][0], value_dict['rpower1'][0], i + 1, Peris_ON,
                                    value_dict['RPM'][0], Iso_ON, IsLog, corrected_time, value_dict['donneestemps'],
                                    value_dict['rpowergraph'], value_dict['powergraph'], Dielec_Verif, value_dict["Listbox2"][0], Timer_ON)
                
                if Exit == 1:
                    break

                if Exit == 2 and i < 6:
                    continue
                
            window.FindElement('_FIVEIT_FRAME_').Update(visible=True)
            sg.popup("Test is complete!")

            try:
                # Calcul des valeurs moyennes automatiser
                value_dict['stdevreel'] = []
                value_dict['stdevim'] = []

                S_Averages(value_dict["directory"][0], value_dict["datapoints"][0], value_dict['donneestemps'], value_dict['stdevreel'], value_dict['stdevim'])

            except:
                sg.popup('Please check the average data fields')
                window.FindElement('_FIVEIT_FRAME_').Update(visible=True)

            # window.reappear()

            # except:
               # sg.popup('Please check the fields')
               # window.FindElement('_FIVEIT_FRAME_').Update(visible=True)

        # Normal test
        if event == '_LAUNCH_OK_':
            try:
                # window.disappear()
                # sg.popup("Press OK to confirm launch. Press X to exit")
                Normal_Test(value_dict['startfreq'][0], value_dict['stopfreq'][0], value_dict['datapoints'][0],
                            value_dict['bw'][0], value_dict['num_its'][0], value_dict['ondelay1'][0],
                            value_dict['offdelay1'][0],
                            value_dict['directory'][0],
                            value_dict['freq1'][0], value_dict['power1'][0], value_dict['rpower1'][0],
                            Peris_ON, value_dict['RPM'][0], Iso_ON, IsLog)
                sg.popup("Test is complete!")
                window.FindElement('_TEST_FRAME_').Update(visible=True)
                window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
                # window.reappear()

            except:
                # window.reappear()
                sg.popup('Please check the fields')

        """
        ALL PARAMETERS WINDOW
        """
        if event == '_ANA_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=True)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_SWITCH_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_GEN_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=True)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_PERIS_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=True)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_ALLPARAMETERS_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=True)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        """
        ANALYZER WINDOW
        """
        if event == '_ANA_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=True)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_FREQ_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=True)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_PARAMS1_OK_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=True)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        """
        PARAMS1 WINDOW
        """
        if event == '_PARAMS1_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=True)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_CONFIG1_OK_':

            value_dict["num_its"] = []
            value_dict["directory"] = []
            try:
                value_dict["num_its"].append(int(values['_NUMITS_']))
                value_dict["directory"].append(str(values['_DIRECTORY_']))

                window.FindElement('_PARAMS1_FRAME_').Update(visible=True)

                sg.popup("Values are set!")

            except:
                sg.popup('Please check the fields')

        """
        FREQ WINDOW
        """
        if event == '_FREQ_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=True)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_CONFIG2_OK_':

            value_dict["startfreq"] = []
            value_dict["stopfreq"] = []
            value_dict["datapoints"] = []
            value_dict["bw"] = []
            # value_dict["inputpower"] = []

            try:
                value_dict["startfreq"].append(float(values['_STARTFREQ_']))
                value_dict["stopfreq"].append(float(values['_STOPFREQ_']))
                value_dict["datapoints"].append(int(values['_NUMDATA_']))
                value_dict["bw"].append(int(values['_BW_']))
                # value_dict["inputpower"].append(int(values['_INPUTPOWER_']))

                window.FindElement('_FREQ_FRAME_').Update(visible=True)

                sg.popup("Values are set!")

            except:
                sg.popup('Please check the fields')

        if event == '_LOG_ON_':
            IsLog = 1

        if event == '_LOG_OFF_':
            IsLog = 0


        """
        GENERATOR WINDOW
        """
        if event == '_GEN_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=True)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_SET_GEN_':

            value_dict["power1"] = []
            value_dict["ondelay1"] = []
            value_dict["offdelay1"] = []
            value_dict["freq1"] = []

            try:

                value_dict["power1"].append(int(values['_POWER1_']))

                # Clause de sécurité. Prévenir l'utilisateur que les valeurs de puissance configurés requis l'accompagnement de la pompe à refroidisement.
                if 20 <= value_dict["power1"][0] < 200:
                    sg.popup(
                        "For the values of transmitted power you've selected, cooling might/will be required.\nPress OK to continue.")

                # Clause de sécurité. Prévenir à ce que l'utilisateur choisisse une valeur de puissance au-delà de 200 W.
                if value_dict["power1"][0] >= 200:
                    sg.popup(
                        "The value of transmitted power selected is too high.\nPlease select another value. The value of transmitted power has been set to 0 Watts.\nPress OK to continue.")
                    value_dict["power1"][0] = 0

                value_dict["ondelay1"].append(float(values['_ONDELAY1_']))
                value_dict["offdelay1"].append(float(values['_OFFDELAY1_']))
                value_dict["freq1"].append(int(values['_FREQ1_']))

                window.FindElement('_GEN_FRAME_').Update(visible=True)

                sg.popup("Values are set!")

            except:
                sg.popup('Please check the fields')

        if event == '_SET_RPOWER_':

            value_dict["rpower1"] = []

            try:
                value_dict["rpower1"].append(float(values['_REFPOWER_']))

                if int(value_dict["rpower1"][0]) != 0:
                    sg.popup(
                        "Are you sure you want to manually configure the reflected power?\nIts value should not exceed 15% of the transmitted powers' value.\nPress OK to continue.")

                window.FindElement('_GEN_FRAME_').Update(visible=True)

                sg.popup("Values are set!")

            except:
                sg.popup('Please check the fields')

        if event == '_RESET_FAULTS_':
            # Reset generator faults function
            Reset_Faults()

        """
        PERISTALTIC PUMP WINDOW
        """
        if event == '_PERIS_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=True)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_CONFIG7_OK_':

            value_dict["RPM"] = []

            try:
                value_dict["RPM"].append(int(values['_RPM_']))

                # Clause de sécurité. Prévenir à ce que l'utilisateur choisisse une valeur de vitesse de débit chez la pompe péristaltique au-delà de 300 RPM.
                if value_dict["RPM"][0] >= 120:
                    sg.popup(
                        "The value of cooling speed selected is too high.\nPlease select another value. The value of speed has been set to 0 RPM.\nPress OK to continue.")
                    value_dict["RPM"][0] = 0

                window.FindElement('_PUMPS_FRAME_').Update(visible=True)

                sg.popup("Values are set!")

            except:
                sg.popup('Please check the fields')

        if event == '_PERIS_ON_':
            Peris_ON = 1

        if event == '_PERIS_OFF_':
            Peris_ON = 0

        if event == '_ISO_ON_':
            Iso_ON = 1

        if event == '_ISO_OFF_':
            Iso_ON = 0

        """
        TEST WINDOW
        """
        # Select button executions

        # Open Frequency sweep test window
        if event == '_TEST_OK1_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=True)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        # Open Manual Generator control window
        if event == '_TEST_OK2_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=True)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        # Open Five iteration test window
        if event == '_TEST_OK3_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=True)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        # Open Manual ENA control window
        if event == '_TEST_OK4_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=True)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        # Open Manual Pumps control window
        if event == '_TEST_OK5_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=True)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        # Open Manual S1P average control window
        if event == '_TEST_OK6_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=True)

        # Cancel button execution
        if event == '_TEST_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=True)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=False)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=False)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        """
        MANUALS, SCAN SWEEP AND FIVE ITERATION TEST WINDOWS
        """
        # Cancel buttons executions
        if event == '_SWEEP_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=True)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_MANUGEN_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=True)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_MANUENA_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=True)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_MANUPUMPS_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=True)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUENA_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_FIVEIT_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=True)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)

        if event == '_MANUS1P_CANCEL_':
            window.FindElement('_INIT_FRAME_').Update(visible=False)
            window.FindElement('_ALLPARAMETERS_FRAME_').Update(visible=False)
            window.FindElement('_GEN_FRAME_').Update(visible=False)
            window.FindElement('_ANA_FRAME_').Update(visible=False)
            window.FindElement('_FREQ_FRAME_').Update(visible=False)
            window.FindElement('_PARAMS1_FRAME_').Update(visible=False)
            window.FindElement('_PUMPS_FRAME_').Update(visible=False)
            window.FindElement('_TEST_FRAME_').Update(visible=True)
            window.FindElement('_SWEEP_FRAME_').Update(visible=False)
            window.FindElement('_MANUGEN_FRAME_').Update(visible=False)
            window.FindElement('_MANUPUMPS_FRAME_').Update(visible=False)
            window.FindElement('_FIVEIT_FRAME_').Update(visible=False)
            window.FindElement('_NORMALTEXT_FRAME_').Update(visible=True)
            window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=False)


        # Sweep scan test configuration
        if event == '_CONFIG4_OK_':

            value_dict["startfreq1"] = []
            value_dict["stopfreq1"] = []
            value_dict["stepfreq"] = []

            try:
                value_dict["startfreq1"].append(int(values['_STARTFREQ1_']))
                value_dict["stopfreq1"].append(int(values['_STOPFREQ1_']))
                value_dict["stepfreq"].append(int(values['_STEPFREQ_']))

                window.FindElement('_SWEEP_FRAME_').Update(visible=True)

                sg.popup("Values are set!")

            except:
                sg.popup('Please check the fields')

        if event == '_MANU_MODE_':
            IsManu = 1

        if event == '_AUTO_MODE_':
            IsManu = 0

        # Five iteration test configuration
        if event == '_CONFIG5_OK_':

            value_dict["itpower"] = []
            value_dict["ondelay3"] = []
            value_dict["Listbox1"] = []
            value_dict["Listbox2"] = []
            value_dict["delaimicro"] = []
            value_dict["delaimesure"] = []
            value_dict["directory"] = []

            try:
                value_dict["itpower"].append(float(values['_ITPOWER1_']))
                value_dict["itpower"].append(float(values['_ITPOWER2_']))
                value_dict["itpower"].append(float(values['_ITPOWER3_']))
                value_dict["itpower"].append(float(values['_ITPOWER4_']))
                value_dict["itpower"].append(float(values['_ITPOWER5_']))

                # Clause de sécurité. Prévenir l'utilisateur que les valeurs de puissance configurés requis l'accompagnement de la pompe à refroidisement.
                if (20 <= value_dict["itpower"][0] < 200) or (20 <= value_dict["itpower"][1] < 200) or (
                        20 <= value_dict["itpower"][2] < 200) or (20 <= value_dict["itpower"][3] < 200) or (
                        20 <= value_dict["itpower"][4] < 200):
                    sg.popup("For the values of transmitted power you've selected, cooling might/will be required.\nPress OK to continue.")

                # Clause de sécurité. Prévenir à ce que l'utilisateur choisisse une valeur de puissance au-delà de 200 W.
                if value_dict["itpower"][0] >= 200 or value_dict["itpower"][1] >= 200 or value_dict["itpower"][
                    2] >= 200 or value_dict["itpower"][3] >= 200 or value_dict["itpower"][4] >= 200:
                    sg.popup("The value of transmitted power selected is too high.\nPlease select another value. All values of transmitted power has been set to 0 Watts.\nPress OK to continue.")
                    value_dict["itpower"][0] = 0
                    value_dict["itpower"][1] = 0
                    value_dict["itpower"][2] = 0
                    value_dict["itpower"][3] = 0
                    value_dict["itpower"][4] = 0

                value_dict["ondelay3"].append(float(values['_ONDELAY31_']))
                value_dict["ondelay3"].append(float(values['_ONDELAY32_']))
                value_dict["ondelay3"].append(float(values['_ONDELAY33_']))
                value_dict["ondelay3"].append(float(values['_ONDELAY34_']))
                value_dict["ondelay3"].append(float(values['_ONDELAY35_']))

                value_dict["Listbox1"].append(float(values['_LISTBOX1_']))
                value_dict["Listbox1"].append(float(values['_LISTBOX2_']))
                value_dict["Listbox1"].append(float(values['_LISTBOX3_']))
                value_dict["Listbox1"].append(float(values['_LISTBOX4_']))
                value_dict["Listbox1"].append(float(values['_LISTBOX5_']))
                value_dict["Listbox2"].append(float(values['_LISTBOX6_']))

                value_dict["delaimesure"].append(float(values['_DELAIMESURE_']))
                value_dict["delaimicro"].append(float(values['_DELAIMICRO_']))
                value_dict["directory"].append(str(values['_DIRECTORY_']))

                window.FindElement('_FIVEIT_FRAME_').Update(visible=True)

                sg.popup("Configured!")

            except:
                sg.popup('Please check the fields')

        # Manual ENA trigger configuration
        if event == '_CONFIG6_OK_':

            value_dict["startfreqmanu"] = []
            value_dict["stopfreqmanu"] = []
            value_dict["datapointsmanu"] = []
            value_dict["bwmanu"] = []
            value_dict["delayENAmanu"] = []
            value_dict["directorymanu"] = []

            try:
                value_dict["startfreqmanu"].append(int(values['_STARTFREQMANU_']))
                value_dict["stopfreqmanu"].append(int(values['_STOPFREQMANU_']))
                value_dict["datapointsmanu"].append(int(values['_DATAPOINTSMANU_']))
                value_dict["bwmanu"].append(int(values['_BWMANU_']))
                value_dict["delayENAmanu"].append(float(values['_DELAYENAMANU_']))
                value_dict["directorymanu"].append(str(values['_DIRECTORYMANU_']))

                window.FindElement('_MANUENA_FRAME_').Update(visible=True)

                sg.popup("Values Set!")

            except:
                sg.popup('Please check the fields')

        # Execute creation of S1P average file
        if event == '_S1PAVERAGE_':

            value_dict['stdevreel'] = []
            value_dict['stdevim'] = []

            try:
                S_Averages(value_dict["directory"][0], value_dict["datapoints"][0], value_dict['donneestemps'], value_dict['stdevreel'], value_dict['stdevim'])

                print(value_dict['stdevreel'])
                sg.popup("File has been created in the folder DonneesMoyennes !")

            except:
                sg.popup('Please check the fields')
                window.FindElement('_FIVEIT_FRAME_').Update(visible=True)

        # Format Excel data file
        if event == '_EXCELDATA_':
            try:
                Excel_Data_Format(value_dict['donneestemps'], value_dict['rpowergraph'], value_dict['powergraph'], value_dict['stdevreel'],value_dict['stdevim'], value_dict["directory"][0])

                sg.popup("Excel file has been formatted!")

            except:
                sg.popup('Please check the fields')
                window.FindElement('_FIVEIT_FRAME_').Update(visible=True)

        # Create average S1P files manually
        if event == '_MANUS1P_OK_':
            try:
                S_Averages_Manu(value_dict['directorymanu1'][0], value_dict['datapointsmanu1'][0], value_dict["num_itsmanu"][0],value_dict['stdevreel'],value_dict['stdevim'])

                sg.popup("S1P averages are now ready!")

            except:
                sg.popup('Please check the fields')
                window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=True)


        # Manual log selection on ENA manual control
        if event == '_LOG_ON1_':
            IsLogManu = 1

        if event == '_LOG_OFF1_':
            IsLogManu = 0

        # Manual timer activation selection
        if event == '_TIMER_ON_':
            Timer_ON = 1

        if event  == '_TIMER_OFF_':
            Timer_ON = 0

        # Manual pumps control
        if event == '_MANU_ISO_ON_':
            Manu_Pumps_State = 1
            Manu_Iso_ON = 1

        if event == '_MANU_ISO_OFF_':
            Manu_Pumps_State = 2
            Manu_Iso_ON = 0

        if event == '_MANU_PERIS_ON_':
            Manu_Pumps_State = 3
            Manu_Peris_ON = 1

        if event == '_MANU_PERIS_OFF_':
            Manu_Pumps_State = 4
            Manu_Peris_ON = 0

        if event == '_MANU_PERIS_OFF_':
            Manu_Pumps_State = 4
            Manu_Peris_ON = 0

        # Dielectric verification selection
        if event == '_DIELEC_VERIF_':
            Dielec_Verif = 1

        if event == '_DIELEC_VERIF1_':
            Dielec_Verif = 0

        # Manual S1P average control
        if event == '_CONFIG8_OK_':

            value_dict["datapointsmanu1"] = []
            value_dict["directorymanu1"] = []
            value_dict["num_itsmanu"] = []

            try:
                value_dict["datapointsmanu1"].append(int(values['_DATAPOINTS_S1P_']))
                value_dict["directorymanu1"].append(str(values['_DIRECTORY_S1P_']))
                value_dict["num_itsmanu"].append(int(values['_ITERATIONS_S1P_']))

                window.FindElement('_S1PAVERAGE_FRAME_').Update(visible=True)

                sg.popup("Values Set!")

            except:
                sg.popup('Please check the fields')


        """
        QUIT WINDOW
        """
        if event is None or event == 'QUIT':

            # Imprimer valeurs de temps et de puissance afin de s'assurer aucune donnee n'est perdue lorsque le code fait faillite
            print(value_dict['donneestemps'])
            print(value_dict['rpowergraph'])
            print(value_dict['powergraph'])

            break


# try:
init()
main()  # Main program

# except:
    # sg.popup("Verify the connections")
    # sys.exit()
